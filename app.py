import os
from flask import Flask, request, jsonify, render_template, redirect, url_for, flash, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import re
from datetime import datetime, timezone, timedelta
import uuid
import zipfile
from io import BytesIO
from collections import defaultdict
from sqlalchemy import func

app = Flask(__name__, template_folder='app/templates', static_folder='app/static')

# 配置
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////app/data/homework.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024 * 1024  # 10GB - 支持教师设置的最大文件大小

# 确保上传目录和附件目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('/app/data', exist_ok=True)
os.makedirs('/app/appendix', exist_ok=True)  # 创建附件目录，使用绝对路径

os.makedirs('appendix', exist_ok=True)  # 创建附件目录

# 初始化扩展
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 北京时区（UTC+8）
BEIJING_TZ = timezone(timedelta(hours=8))

def safe_chinese_filename(filename):
    """
    创建支持中文的安全文件名
    保留中文字符、英文字符、数字，过滤危险字符
    """
    if not filename:
        return 'untitled'
    
    # 移除或替换危险字符，但保留中文字符
    # 保留：中文字符、英文字母、数字、短横线、下划线、圆括号、方括号、空格
    safe_name = re.sub(r'[<>:"/\\|?*]', '', filename)  # 移除文件系统危险字符
    safe_name = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', safe_name)  # 移除控制字符
    safe_name = safe_name.strip()  # 移除首尾空格
    
    # 如果文件名为空或只有空格，使用默认名称
    if not safe_name:
        return 'untitled'
    
    # 限制长度（考虑到中文字符）
    if len(safe_name.encode('utf-8')) > 200:
        # 截断但确保不会在中文字符中间截断
        truncated = safe_name[:100]  # 先截断到100个字符
        while len(truncated.encode('utf-8')) > 200 and len(truncated) > 0:
            truncated = truncated[:-1]
        safe_name = truncated
    
    return safe_name

def to_beijing_time(utc_dt):
    """将UTC时间转换为北京时间"""
    if utc_dt is None:
        return None
    if utc_dt.tzinfo is None:
        # 假设是UTC时间
        utc_dt = utc_dt.replace(tzinfo=timezone.utc)
    return utc_dt.astimezone(BEIJING_TZ)

# 注册Jinja2过滤器
@app.template_filter('beijing_time')
def beijing_time_filter(utc_dt):
    """模板过滤器：将UTC时间转换为北京时间并格式化"""
    beijing_dt = to_beijing_time(utc_dt)
    if beijing_dt is None:
        return '未知'
    return beijing_dt.strftime('%Y-%m-%d %H:%M:%S')

@app.template_filter('beijing_date')
def beijing_date_filter(utc_dt):
    """模板过滤器：将UTC时间转换为北京时间日期"""
    beijing_dt = to_beijing_time(utc_dt)
    if beijing_dt is None:
        return '未知'
    return beijing_dt.strftime('%Y-%m-%d')

@app.template_filter('beijing_short')
def beijing_short_filter(utc_dt):
    """模板过滤器：将UTC时间转换为北京时间短格式"""
    beijing_dt = to_beijing_time(utc_dt)
    if beijing_dt is None:
        return '未知'
    return beijing_dt.strftime('%m-%d %H:%M')

@app.template_filter('beijing_datetime_local')
def beijing_datetime_local_filter(utc_dt):
    """模板过滤器：将UTC时间转换为北京时间，用于datetime-local输入框"""
    beijing_dt = to_beijing_time(utc_dt)
    if beijing_dt is None:
        return ''
    return beijing_dt.strftime('%Y-%m-%dT%H:%M')

@app.template_filter('nl2br')
def nl2br_filter(text):
    """模板过滤器：将换行符转换为HTML换行标签"""
    if not text:
        return ''
    from markupsafe import Markup
    return Markup(text.replace('\n', '<br>'))

# 用户角色枚举
class UserRole:
    SUPER_ADMIN = 'super_admin'  # 超级管理员
    TEACHER = 'teacher'          # 教师
    STUDENT = 'student'          # 学生

# 班级模型
class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 班级名称
    code = db.Column(db.String(50), unique=True, nullable=False)  # 班级代码
    description = db.Column(db.Text)  # 班级描述
    grade = db.Column(db.String(50))  # 年级
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # 创建者
    
    # 关系
    creator = db.relationship('User', backref='created_classes', foreign_keys=[created_by])
    students = db.relationship('User', secondary='class_student', backref='classes')
    teachers = db.relationship('User', secondary='class_teacher', backref='teaching_classes')
    
    def __repr__(self):
        return f'<Class {self.name}>'

# 班级-学生关联表
class_student = db.Table('class_student',
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True),
    db.Column('student_id', db.Integer, db.ForeignKey('user.id'), primary_key=True)
)

# 班级-教师关联表  
class_teacher = db.Table('class_teacher',
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True),
    db.Column('teacher_id', db.Integer, db.ForeignKey('user.id'), primary_key=True)
)

# 用户模型 - 扩展版本
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)  # 可以是用户名或真实姓名
    real_name = db.Column(db.String(100), nullable=False)  # 真实姓名
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default=UserRole.STUDENT)
    student_id = db.Column(db.String(50), unique=True)  # 学号，学生专用
    is_active = db.Column(db.Boolean, default=True)
    must_change_password = db.Column(db.Boolean, default=True)  # 首次登录必须修改密码
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # 创建者
    
    # 关系
    created_users = db.relationship('User', backref=db.backref('creator', remote_side=[id]))
    assignments = db.relationship('Assignment', backref='teacher', lazy=True)
    submissions = db.relationship('Submission', backref='student_user', lazy=True, foreign_keys='Submission.student_id')
    graded_submissions = db.relationship('Submission', backref='grader', lazy=True, foreign_keys='Submission.graded_by')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    @property
    def is_super_admin(self):
        return self.role == UserRole.SUPER_ADMIN
    
    @property
    def is_teacher(self):
        return self.role == UserRole.TEACHER
    
    @property
    def is_student(self):
        return self.role == UserRole.STUDENT
    
    def can_manage_users(self):
        return self.role in [UserRole.SUPER_ADMIN, UserRole.TEACHER]
    
    def can_create_assignments(self):
        return self.role in [UserRole.SUPER_ADMIN, UserRole.TEACHER]
    
    def can_reset_system(self):
        return self.role == UserRole.SUPER_ADMIN

# 作业模型 - 扩展版本
class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    due_date = db.Column(db.DateTime)  # 截止时间
    allowed_file_types = db.Column(db.Text)  # 允许的文件类型，用逗号分隔
    max_file_size = db.Column(db.Integer, default=50*1024*1024)  # 最大文件大小，默认50MB
    max_submissions = db.Column(db.Integer, default=1)  # 最大提交次数，默认1次
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'))  # 所属班级
    is_active = db.Column(db.Boolean, default=True)
    
    # 附件相关字段
    attachment_filename = db.Column(db.String(255))  # 附件文件名
    attachment_original_filename = db.Column(db.String(255))  # 附件原始文件名
    attachment_file_path = db.Column(db.String(500))  # 附件文件路径
    attachment_file_size = db.Column(db.Integer)  # 附件文件大小
    
    # 关系
    submissions = db.relationship('Submission', backref='assignment', lazy=True, cascade='all, delete-orphan')
    class_info = db.relationship('Class', backref='assignments')
    
    def get_allowed_extensions(self):
        """获取允许的文件扩展名列表"""
        if self.allowed_file_types:
            return [ext.strip().lower() for ext in self.allowed_file_types.split(',')]
        return ['pdf', 'doc', 'docx', 'txt', 'zip', 'rar']  # 默认允许的类型
    
    def is_file_allowed(self, filename):
        """检查文件类型是否允许"""
        if '.' not in filename:
            return False
        ext = filename.rsplit('.', 1)[1].lower()
        return ext in self.get_allowed_extensions()
    
    def is_overdue(self):
        """检查是否已过截止时间"""
        if not self.due_date:
            return False
        return datetime.utcnow() > self.due_date
    
    def get_student_submission_count(self, student_id):
        """获取学生提交次数"""
        return Submission.query.filter_by(assignment_id=self.id, student_id=student_id).count()
    
    def can_student_submit(self, student_id):
        """检查学生是否还能提交作业"""
        if self.max_submissions <= 0:  # 无限制提交
            return True
        return self.get_student_submission_count(student_id) < self.max_submissions

# 作业评分模型 - 用于存储多教师对同一作业的评分
class AssignmentGrade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 评分教师
    grade = db.Column(db.Float)  # 成绩
    feedback = db.Column(db.Text)  # 教师反馈
    graded_at = db.Column(db.DateTime, default=datetime.utcnow)  # 评分时间
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)  # 更新时间
    
    # 关系
    assignment = db.relationship('Assignment', backref='assignment_grades')
    student = db.relationship('User', foreign_keys=[student_id], backref='received_grades')
    teacher = db.relationship('User', foreign_keys=[teacher_id], backref='given_grades')
    
    # 复合唯一约束：同一教师对同一学生的同一作业只能有一个评分
    __table_args__ = (db.UniqueConstraint('assignment_id', 'student_id', 'teacher_id', name='unique_assignment_student_teacher_grade'),)

# 提交模型 - 扩展版本
class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # 关联到用户表
    student_name = db.Column(db.String(100), nullable=False)  # 冗余字段，保持兼容性
    student_number = db.Column(db.String(50), nullable=False)  # 学号
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_size = db.Column(db.Integer)  # 文件大小
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)
    grade = db.Column(db.Float)  # 成绩
    feedback = db.Column(db.Text)  # 教师反馈
    graded_at = db.Column(db.DateTime)  # 评分时间
    graded_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # 评分教师
    
    def is_pdf(self):
        """检查文件是否为PDF格式"""
        return self.original_filename.lower().endswith('.pdf')
    
    def get_file_url(self):
        """获取文件访问URL"""
        return url_for('download_file', submission_id=self.id)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# 权限装饰器
def require_role(role):
    """要求特定角色的装饰器"""
    def decorator(f):
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role != role and current_user.role != UserRole.SUPER_ADMIN:
                flash('您没有权限访问此页面')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        decorated_function.__name__ = f.__name__
        return decorated_function
    return decorator

def require_teacher_or_admin(f):
    """要求教师或管理员权限"""
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not (current_user.is_teacher or current_user.is_super_admin):
            flash('您没有权限访问此页面')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# 文件处理函数
def allowed_file(filename, allowed_extensions=None):
    """检查文件类型是否允许"""
    if not allowed_extensions:
        allowed_extensions = {'txt', 'pdf', 'doc', 'docx', 'zip', 'rar', 'py', 'java', 'cpp', 'c', 'html', 'css', 'js'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def get_file_size(file_path):
    """获取文件大小"""
    try:
        return os.path.getsize(file_path)
    except:
        return 0

def save_assignment_attachment(attachment_file):
    """保存作业附件"""
    print(f"DEBUG: save_assignment_attachment called with file: {attachment_file}")
    if not attachment_file or not attachment_file.filename:
        print("DEBUG: No attachment file provided")
        return None, None, None, None
    
    # 生成安全的文件名
    original_filename = attachment_file.filename
    filename = secure_filename(original_filename)
    print(f"DEBUG: Original filename: {original_filename}, Secure filename: {filename}")
    
    if not filename:
        # 如果secure_filename返回空字符串（例如全中文文件名），则生成一个随机文件名
        file_ext = os.path.splitext(original_filename)[1] if '.' in original_filename else ''
        filename = str(uuid.uuid4()) + file_ext
        print(f"DEBUG: Generated random filename: {filename}")
    
    # 生成唯一的文件路径（使用绝对路径）
    unique_filename = f"{uuid.uuid4()}_{filename}"
    file_path = os.path.join('/app/appendix', unique_filename)
    print(f"DEBUG: File path: {file_path}")
    
    try:
        # 确保目录存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        # 保存文件
        attachment_file.save(file_path)
        print(f"DEBUG: File saved successfully")
        
        # 获取文件大小
        file_size = get_file_size(file_path)
        print(f"DEBUG: File size: {file_size}")
        
        return filename, original_filename, file_path, file_size
    except Exception as e:
        print(f"DEBUG: Error saving file: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None, None

def delete_assignment_attachment(file_path):
    """删除作业附件"""
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
            return True
        except Exception as e:
            print(f"删除附件失败: {e}")
            return False
    return True

# 路由
@app.route('/')
def index():
    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    return render_template('index.html', assignments=assignments)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # 先尝试用户名登录
        user = User.query.filter_by(username=username).first()
        
        # 如果用户名没找到，尝试用真实姓名登录
        if not user:
            user = User.query.filter_by(real_name=username).first()
        
        # 如果还没找到，尝试用学号或教工号登录（仅对学生和教师）
        if not user:
            user = User.query.filter_by(student_id=username).first()
        
        if user and user.check_password(password) and user.is_active:
            login_user(user)
            
            # 检查是否需要强制修改密码（超级管理员除外）
            if user.must_change_password and not user.is_super_admin:
                flash('您是首次登录，必须修改密码后才能继续使用系统')
                return redirect(url_for('force_change_password'))
            
            # 根据角色重定向到不同的页面
            if user.is_super_admin:
                return redirect(url_for('super_admin_dashboard'))
            elif user.is_teacher:
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            flash('用户名或密码错误，或者账户已被禁用')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# 测试附件上传功能
@app.route('/test-attachment')
def test_attachment():
    """测试附件上传功能"""
    from werkzeug.datastructures import FileStorage
    from io import BytesIO
    import os
    
    # 创建一个测试附件文件
    test_content = b'This is a test attachment file for homework assignment.'
    test_file = FileStorage(
        stream=BytesIO(test_content),
        filename='test_attachment.txt',
        content_type='text/plain'
    )
    
    # 直接调用保存附件函数
    filename, original_filename, file_path, file_size = save_assignment_attachment(test_file)
    
    result = f"Attachment saved:\n"
    result += f"  filename: {filename}\n"
    result += f"  original_filename: {original_filename}\n"
    result += f"  file_path: {file_path}\n"
    result += f"  file_size: {file_size}\n"
    
    # 检查文件是否真的存在
    if file_path and os.path.exists(file_path):
        result += f"File exists at {file_path}\n"
        # 读取文件内容验证
        with open(file_path, 'rb') as f:
            content = f.read()
            result += f"File content length: {len(content)}\n"
    else:
        result += "File was not saved correctly\n"
    
    return result

# 测试附件上传功能
@app.route('/test-attachment-upload', methods=['GET', 'POST'])
def test_attachment_upload():
    """测试附件上传功能"""
    if request.method == 'POST':
        print("DEBUG: test_attachment_upload POST request received")
        print(f"DEBUG: request.files: {list(request.files.keys())}")
        
        if 'attachment' not in request.files:
            return "No attachment file in request"
        
        attachment_file = request.files['attachment']
        print(f"DEBUG: attachment_file: {attachment_file}")
        print(f"DEBUG: attachment_file.filename: {attachment_file.filename}")
        
        if not attachment_file or not attachment_file.filename:
            return "No file selected"
        
        # 调用保存附件函数
        filename, original_filename, file_path, file_size = save_assignment_attachment(attachment_file)
        
        result = f"Attachment saved:\n"
        result += f"  filename: {filename}\n"
        result += f"  original_filename: {original_filename}\n"
        result += f"  file_path: {file_path}\n"
        result += f"  file_size: {file_size}\n"
        
        # 检查文件是否真的存在
        if file_path and os.path.exists(file_path):
            result += f"File exists at {file_path}\n"
            # 读取文件内容验证
            with open(file_path, 'rb') as f:
                content = f.read()
                result += f"File content length: {len(content)}\n"
        else:
            result += "File was not saved correctly\n"
            result += f"appendix directory exists: {os.path.exists('appendix')}\n"
            if os.path.exists('appendix'):
                result += f"appendix directory contents: {os.listdir('appendix')}\n"
        
        return result
    
    # GET请求显示上传表单
    return '''
    <!doctype html>
    <title>Upload new File</title>
    <h1>Upload new File</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=attachment>
      <input type=submit value=Upload>
    </form>
    '''

# 强制修改密码路由

# 强制修改密码路由
@app.route('/force-change-password', methods=['GET', 'POST'])
@login_required
def force_change_password():
    # 如果用户不需要强制修改密码，重定向到主页
    if not current_user.must_change_password or current_user.is_super_admin:
        if current_user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        elif current_user.is_teacher:
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # 验证当前密码
        if not current_user.check_password(current_password):
            flash('当前密码错误')
            return render_template('force_change_password.html')
        
        # 验证新密码
        if len(new_password) < 6:
            flash('新密码长度至少6位')
            return render_template('force_change_password.html')
        
        if new_password != confirm_password:
            flash('两次输入的新密码不一致')
            return render_template('force_change_password.html')
        
        if new_password == current_password:
            flash('新密码不能与当前密码相同')
            return render_template('force_change_password.html')
        
        # 更新密码并标记不再需要强制修改
        current_user.set_password(new_password)
        current_user.must_change_password = False
        db.session.commit()
        
        flash('密码修改成功，欢迎使用系统！')
        
        # 根据角色重定向
        if current_user.is_teacher:
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    
    return render_template('force_change_password.html')

# 超级管理员仪表板
@app.route('/super-admin')
@login_required
@require_role(UserRole.SUPER_ADMIN)
def super_admin_dashboard():
    total_users = User.query.count()
    total_teachers = User.query.filter_by(role=UserRole.TEACHER).count()
    total_students = User.query.filter_by(role=UserRole.STUDENT).count()
    total_assignments = Assignment.query.count()
    total_submissions = Submission.query.count()
    total_classes = Class.query.count()
    
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    recent_assignments = Assignment.query.order_by(Assignment.created_at.desc()).limit(5).all()
    
    # 获取所有作业用于管理
    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    
    return render_template('super_admin_dashboard.html', 
                         total_users=total_users,
                         total_teachers=total_teachers,
                         total_students=total_students,
                         total_assignments=total_assignments,
                         total_submissions=total_submissions,
                         total_classes=total_classes,
                         recent_users=recent_users,
                         recent_assignments=recent_assignments,
                         assignments=assignments)

# 教师仪表板
@app.route('/teacher')
@login_required
@require_teacher_or_admin
def teacher_dashboard():
    # 教师可以看到两种作业：
    # 1. 自己创建的作业
    # 2. 分配给自己负责班级的作业（由超级管理员创建）
    
    # 获取教师自己创建的作业
    own_assignments = Assignment.query.filter_by(teacher_id=current_user.id).all()
    
    # 获取教师负责班级的所有作业（包括超级管理员创建的）
    teacher_classes = current_user.teaching_classes
    class_assignments = []
    if teacher_classes:
        class_ids = [c.id for c in teacher_classes]
        class_assignments = Assignment.query.filter(
            Assignment.class_id.in_(class_ids),
            Assignment.teacher_id != current_user.id  # 排除自己已经创建的作业，避免重复
        ).all()
    
    # 合并所有作业并按创建时间排序
    all_assignments = own_assignments + class_assignments
    assignments = sorted(all_assignments, key=lambda x: x.created_at, reverse=True)
    
    # 获取教师创建的学生
    my_students = User.query.filter_by(role=UserRole.STUDENT, created_by=current_user.id).all()
    
    return render_template('teacher_dashboard.html', 
                         assignments=assignments,
                         my_students=my_students)

# 学生仪表板
@app.route('/student')
@login_required
@require_role(UserRole.STUDENT)
def student_dashboard():
    # 获取学生所在班级的作业
    student_classes = current_user.classes
    if student_classes:
        class_ids = [c.id for c in student_classes]
        # 获取学生所在班级的活跃作业
        assignments = Assignment.query.filter(
            Assignment.class_id.in_(class_ids),
            Assignment.is_active == True
        ).order_by(Assignment.created_at.desc()).all()
        
        # 也包含未指定班级的公共作业（兼容性）
        public_assignments = Assignment.query.filter(
            Assignment.class_id.is_(None),
            Assignment.is_active == True
        ).order_by(Assignment.created_at.desc()).all()
        
        assignments.extend(public_assignments)
        # 按创建时间重新排序
        assignments.sort(key=lambda x: x.created_at, reverse=True)
    else:
        # 如果学生没有分配到任何班级，只显示公共作业
        assignments = Assignment.query.filter(
            Assignment.class_id.is_(None),
            Assignment.is_active == True
        ).order_by(Assignment.created_at.desc()).all()
    
    # 获取用户的提交记录
    my_submissions = Submission.query.filter_by(student_id=current_user.id).order_by(Submission.submitted_at.desc()).all()
    
    return render_template('student_dashboard.html', 
                         assignments=assignments,
                         my_submissions=my_submissions)

# 兼容旧的admin路由
@app.route('/admin')
@login_required
def admin():
    if current_user.is_super_admin:
        return redirect(url_for('super_admin_dashboard'))
    elif current_user.is_teacher:
        return redirect(url_for('teacher_dashboard'))
    else:
        return redirect(url_for('student_dashboard'))

# 用户管理路由
@app.route('/admin/users')
@login_required
@require_teacher_or_admin
def manage_users():
    # 根据用户角色显示不同的用户列表
    if current_user.is_super_admin:
        users = User.query.order_by(User.created_at.desc()).all()
    else:
        # 教师可以看到：1.自己创建的学生  2.自己负责班级中的所有学生
        created_students = User.query.filter_by(created_by=current_user.id, role=UserRole.STUDENT).all()
        
        # 获取教师负责班级中的所有学生
        class_students = []
        for class_obj in current_user.teaching_classes:
            class_students.extend(class_obj.students)
        
        # 合并并去重
        all_students = list({user.id: user for user in (created_students + class_students)}.values())
        
        # 按创建时间排序
        users = sorted(all_students, key=lambda x: x.created_at, reverse=True)
    
    # 创建一个辅助函数来检查教师是否有权限管理学生
    def can_teacher_manage_student(teacher, student):
        # 只能管理学生角色
        if student.role != UserRole.STUDENT:
            return False
        
        # 超级管理员可以管理所有学生
        if teacher.is_super_admin:
            return True
        
        # 教师只能管理自己班级中的学生
        if teacher.is_teacher:
            # 检查学生是否在教师负责的班级中
            for class_obj in teacher.teaching_classes:
                if student in class_obj.students:
                    return True
        
        return False
    
    return render_template('manage_users.html', users=users, can_teacher_manage_student=can_teacher_manage_student)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def add_user():
    if request.method == 'POST':
        username = request.form['username']
        real_name = request.form['real_name']
        password = request.form['password']
        role = request.form['role']
        student_id = request.form.get('student_id', '')
        class_ids = request.form.getlist('classes')
        
        # 确保真实姓名和用户名一致
        username = real_name
        
        # 权限检查
        if not current_user.is_super_admin and role != UserRole.STUDENT:
            flash('您只能添加学生用户')
            # 获取可选班级列表
            if current_user.is_super_admin:
                available_classes = Class.query.filter_by(is_active=True).all()
            else:
                available_classes = current_user.teaching_classes
            return render_template('add_user.html', available_classes=available_classes)
        
        # 检查用户名是否已存在
        if User.query.filter_by(username=username).first():
            flash('用户名已存在')
            # 获取可选班级列表
            if current_user.is_super_admin:
                available_classes = Class.query.filter_by(is_active=True).all()
            else:
                available_classes = current_user.teaching_classes
            return render_template('add_user.html', available_classes=available_classes)
        
        # 检查真实姓名是否已存在
        if User.query.filter_by(real_name=real_name).first():
            flash('真实姓名已存在')
            # 获取可选班级列表
            if current_user.is_super_admin:
                available_classes = Class.query.filter_by(is_active=True).all()
            else:
                available_classes = current_user.teaching_classes
            return render_template('add_user.html', available_classes=available_classes)
        
        # 如果是学生，检查学号是否已存在（只有非空学号才检查）
        if role == UserRole.STUDENT and student_id.strip():
            if User.query.filter_by(student_id=student_id.strip()).first():
                flash('学号已存在')
                # 获取可选班级列表
                if current_user.is_super_admin:
                    available_classes = Class.query.filter_by(is_active=True).all()
                else:
                    available_classes = current_user.teaching_classes
                return render_template('add_user.html', available_classes=available_classes)
        
        # 创建新用户
        user = User(
            username=username,
            real_name=real_name,
            role=role,
            student_id=student_id if (role == UserRole.STUDENT and student_id.strip()) else None,
            created_by=current_user.id,
            must_change_password=False if role == UserRole.SUPER_ADMIN else True  # 超级管理员不需要强制修改密码
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.flush()  # 获取用户ID
        
        # 处理班级分配
        if role == UserRole.TEACHER and current_user.is_super_admin:
            # 教师班级分配
            teacher_class_ids = request.form.getlist('teacher_classes')
            if teacher_class_ids:
                classes_to_assign = Class.query.filter(Class.id.in_(teacher_class_ids)).all()
                user.teaching_classes.extend(classes_to_assign)
        elif role == UserRole.STUDENT:
            # 学生班级分配
            if class_ids:
                if current_user.is_super_admin:
                    # 超级管理员可以分配到任意班级
                    classes = Class.query.filter(Class.id.in_(class_ids)).all()
                else:
                    # 教师只能分配到自己的班级
                    teacher_class_ids = [str(c.id) for c in current_user.teaching_classes]
                    valid_class_ids = [cid for cid in class_ids if cid in teacher_class_ids]
                    classes = Class.query.filter(Class.id.in_(valid_class_ids)).all()
                
                user.classes.extend(classes)
        
        db.session.commit()
        
        flash(f'用户 {real_name} 创建成功')
        return redirect(url_for('manage_users'))
    
    # 获取可选班级列表
    if current_user.is_super_admin:
        available_classes = Class.query.filter_by(is_active=True).all()
    else:
        available_classes = current_user.teaching_classes
    
    return render_template('add_user.html', available_classes=available_classes)

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        # 教师可以编辑自己班级中的学生
        if user.role != UserRole.STUDENT:
            flash('您没有权限编辑此用户')
            return redirect(url_for('manage_users'))
        
        # 检查学生是否在教师的班级中
        student_in_teacher_classes = False
        for class_obj in current_user.teaching_classes:
            if user in class_obj.students:
                student_in_teacher_classes = True
                break
        
        if not student_in_teacher_classes:
            flash('您没有权限编辑此用户')
            return redirect(url_for('manage_users'))
    
    if request.method == 'POST':
        user.real_name = request.form['real_name']
        # 确保用户名和真实姓名一致
        user.username = user.real_name
        user.student_id = request.form.get('student_id', '').strip() if (user.role == UserRole.STUDENT and request.form.get('student_id', '').strip()) else None
        user.is_active = 'is_active' in request.form
        
        # 如果提供了新密码，则更新密码
        new_password = request.form.get('password')
        if new_password:
            if len(new_password) < 6:
                flash('密码长度至少6位')
                # 获取班级列表
                if current_user.is_super_admin:
                    available_classes = Class.query.filter_by(is_active=True).all()
                else:
                    available_classes = current_user.teaching_classes
                return render_template('edit_user.html', user=user, available_classes=available_classes)
            
            user.set_password(new_password)
            # 如果是超级管理员修改其他用户密码，标记为需要强制修改密码（除非是超级管理员）
            if current_user.is_super_admin and user.id != current_user.id and not user.is_super_admin:
                user.must_change_password = True
                flash(f'用户 {user.real_name} 的密码已更新，该用户下次登录时必须修改密码')
            else:
                flash(f'用户 {user.real_name} 的密码已更新')
        
        # 只有超级管理员可以修改角色
        old_role = user.role
        if current_user.is_super_admin:
            new_role = request.form['role']
            user.role = new_role
            
            # 如果角色发生变化，清空班级关联
            if old_role != new_role:
                if old_role == UserRole.TEACHER:
                    user.teaching_classes.clear()
                elif old_role == UserRole.STUDENT:
                    user.classes.clear()
            
            # 处理班级分配
            if new_role == UserRole.TEACHER:
                # 教师班级分配
                teacher_class_ids = request.form.getlist('teacher_classes')
                user.teaching_classes.clear()
                if teacher_class_ids:
                    classes_to_assign = Class.query.filter(Class.id.in_(teacher_class_ids)).all()
                    user.teaching_classes.extend(classes_to_assign)
            elif new_role == UserRole.STUDENT:
                # 学生班级分配
                student_class_ids = request.form.getlist('student_classes')
                user.classes.clear()
                if student_class_ids:
                    classes_to_assign = Class.query.filter(Class.id.in_(student_class_ids)).all()
                    user.classes.extend(classes_to_assign)
        elif user.role == UserRole.STUDENT:
            # 教师可以修改学生的班级分配
            student_class_ids = request.form.getlist('student_classes')
            # 清空并重新分配
            user.classes.clear()
            if student_class_ids:
                # 教师只能分配到自己负责的班级
                teacher_class_ids = [str(c.id) for c in current_user.teaching_classes]
                valid_class_ids = [cid for cid in student_class_ids if cid in teacher_class_ids]
                classes_to_assign = Class.query.filter(Class.id.in_(valid_class_ids)).all()
                user.classes.extend(classes_to_assign)
        
        db.session.commit()
        flash(f'用户 {user.real_name} 信息已更新')
        return redirect(url_for('manage_users'))
    
    # GET请求：获取班级列表
    if current_user.is_super_admin:
        available_classes = Class.query.filter_by(is_active=True).all()
    else:
        available_classes = current_user.teaching_classes
    
    return render_template('edit_user.html', user=user, available_classes=available_classes)

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@require_teacher_or_admin
def reset_user_password(user_id):
    """教师或超级管理员重置用户密码"""
    user = User.query.get_or_404(user_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        # 教师只能重置自己班级中的学生的密码
        if user.role != UserRole.STUDENT:
            flash('您没有权限重置此用户的密码')
            return redirect(url_for('manage_users'))
        
        # 检查学生是否在教师的班级中
        # 获取教师负责的所有班级
        teacher_classes = current_user.teaching_classes
        
        # 检查学生是否在这些班级中
        student_in_teacher_classes = False
        for class_obj in teacher_classes:
            if user in class_obj.students:
                student_in_teacher_classes = True
                break
        
        if not student_in_teacher_classes:
            flash('您没有权限重置此用户的密码')
            return redirect(url_for('manage_users'))
    
    # 不能重置自己的密码
    if user.id == current_user.id:
        flash('不能重置自己的密码，请使用修改密码功能')
        return redirect(url_for('manage_users'))
    
    # 生成默认密码（可以是123456或者用户名等）
    default_password = '123456'  # 可以根据需要修改默认密码规则
    
    user.set_password(default_password)
    # 标记需要强制修改密码（除非是超级管理员）
    if not user.is_super_admin:
        user.must_change_password = True
    
    db.session.commit()
    
    flash(f'用户 {user.real_name} 的密码已重置为默认密码：{default_password}，该用户下次登录时必须修改密码')
    return redirect(url_for('manage_users'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@require_teacher_or_admin
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        # 教师可以删除自己班级中的学生
        if user.role != UserRole.STUDENT:
            flash('您没有权限删除此用户')
            return redirect(url_for('manage_users'))
        
        # 检查学生是否在教师的班级中
        student_in_teacher_classes = False
        for class_obj in current_user.teaching_classes:
            if user in class_obj.students:
                student_in_teacher_classes = True
                break
        
        if not student_in_teacher_classes:
            flash('您没有权限删除此用户')
            return redirect(url_for('manage_users'))
    
    # 不能删除自己
    if user.id == current_user.id:
        flash('不能删除自己的账户')
        return redirect(url_for('manage_users'))
    
    real_name = user.real_name
    db.session.delete(user)
    db.session.commit()
    
    flash(f'用户 {real_name} 已删除')
    return redirect(url_for('manage_users'))

def validate_class_associations():
    """验证班级关联的数据一致性"""
    import sqlite3
    
    with app.app_context():
        conn = sqlite3.connect('/app/data/homework.db')
        cursor = conn.cursor()
        
        # 检查学生表中是否有非学生
        cursor.execute('''
            SELECT COUNT(*) FROM class_student cs
            JOIN user u ON cs.student_id = u.id
            WHERE u.role != 'student'
        ''')
        wrong_students = cursor.fetchone()[0]
        
        # 检查教师表中是否有非教师
        cursor.execute('''
            SELECT COUNT(*) FROM class_teacher ct
            JOIN user u ON ct.teacher_id = u.id
            WHERE u.role != 'teacher'
        ''')
        wrong_teachers = cursor.fetchone()[0]
        
        conn.close()
        
        if wrong_students > 0 or wrong_teachers > 0:
            print(f"⚠️ 数据一致性问题: 学生表中错误记录{wrong_students}个，教师表中错误记录{wrong_teachers}个")
            return False
        else:
            print("✅ 班级关联数据一致性检查通过")
            return True

@app.route('/admin/users/batch-import', methods=['POST'])
@login_required
@require_role(UserRole.SUPER_ADMIN)
def batch_import_users():
    """批量导入用户"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})
    
    if not (file.filename.endswith('.csv') or file.filename.endswith('.tsv') or file.filename.endswith('.txt')):
        return jsonify({'success': False, 'message': '只支持CSV/TSV/TXT文件'})
    
    try:
        # 读取文件，使用更健壮的编码处理
        file_content = file.stream.read()
        
        # 尝试不同的编码方式解码
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'cp1252']
        decoded_content = None
        used_encoding = None
        
        for encoding in encodings:
            try:
                decoded_content = file_content.decode(encoding)
                used_encoding = encoding
                print(f"成功使用编码: {encoding}")
                break
            except UnicodeDecodeError as e:
                print(f"编码 {encoding} 失败: {e}")
                continue
        
        if decoded_content is None:
            return jsonify({'success': False, 'message': '无法解码文件，请确保文件编码正确(支持UTF-8, GBK, GB2312等)'})
        
        import csv
        from io import StringIO
        
        # 使用解码后的内容创建StringIO对象
        csv_data = StringIO(decoded_content)
        
        # 检测分隔符：优先使用制表符，然后是逗号
        sample = decoded_content[:1024]  # 取前1024个字符作为样本
        if '\t' in sample:
            delimiter = '\t'
            print("检测到制表符分隔")
        else:
            delimiter = ','
            print("使用逗号分隔")
        
        try:
            reader = csv.DictReader(csv_data, delimiter=delimiter)
            # 验证CSV格式
            fieldnames = reader.fieldnames
            print(f"CSV字段: {fieldnames}")
            
            if not fieldnames:
                return jsonify({'success': False, 'message': 'CSV文件格式不正确，请检查文件内容'})
            
            # 检查必要字段
            required_fields = ['姓名', '密码', '用户类型']
            missing_fields = [field for field in required_fields if field not in fieldnames]
            if missing_fields:
                return jsonify({
                    'success': False, 
                    'message': f'CSV文件缺少必要字段: {", ".join(missing_fields)}。必要字段包括: 姓名, 密码, 用户类型'
                })
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'CSV文件格式错误: {str(e)}'})
        
        # 重新创建reader来处理数据
        csv_data = StringIO(decoded_content)
        reader = csv.DictReader(csv_data, delimiter=delimiter)
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 开始事务处理
        try:
            # 缓存已创建的班级，避免重复查询
            created_classes = {}
            
            # 使用session.no_autoflush防止自动刷新导致的问题
            with db.session.no_autoflush:
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # 获取并清理数据
                        real_name = row.get('姓名', '').strip()
                        password = row.get('密码', '').strip()
                        class_name = row.get('班级', '').strip()
                        role = row.get('用户类型', '').strip().lower()
                        id_number = row.get('学号/教工号', '').strip()
                        
                        print(f"处理第{row_num}行: {real_name}, {role}, {class_name}")
                        
                        # 验证必要字段
                        if not real_name or not password or not role:
                            errors.append(f'第{row_num}行: 缺少必要字段(姓名、密码、用户类型)')
                            error_count += 1
                            continue
                        
                        # 验证用户类型
                        if role not in ['student', 'teacher']:
                            errors.append(f'第{row_num}行: 无效的用户类型 "{role}"，应为 "student" 或 "teacher"')
                            error_count += 1
                            continue
                        
                        # 检查用户名是否已存在（通过用户名）
                        if User.query.filter_by(username=real_name).first():
                            errors.append(f'第{row_num}行: 用户名 "{real_name}" 已存在')
                            error_count += 1
                            continue
                        
                        # 检查真实姓名是否已存在
                        if User.query.filter_by(real_name=real_name).first():
                            errors.append(f'第{row_num}行: 用户 "{real_name}" 已存在')
                            error_count += 1
                            continue
                        
                        # 检查学号/教工号是否已存在（如果提供了）
                        if id_number and User.query.filter_by(student_id=id_number).first():
                            errors.append(f'第{row_num}行: 学号/教工号 "{id_number}" 已存在')
                            error_count += 1
                            continue
                        
                        # 创建用户
                        user = User(
                            username=real_name,  # 用户名和真实姓名相同
                            real_name=real_name,
                            role=role,
                            student_id=id_number if id_number else None,
                            created_by=current_user.id
                        )
                        user.set_password(password)
                        
                        db.session.add(user)
                        print(f"创建用户: {real_name}")
                        
                        # 处理班级（在缓存中查找或创建）
                        class_obj = None
                        if class_name:
                            if class_name in created_classes:
                                # 从缓存中获取
                                class_obj = created_classes[class_name]
                                print(f"从缓存中获取班级: {class_name}")
                            else:
                                # 查找数据库中的班级
                                class_obj = Class.query.filter_by(name=class_name).first()
                                if not class_obj:
                                    # 创建新班级
                                    import time
                                    class_code = f"C{int(time.time()*1000) % 1000000}"  # 使用时间戳生成唯一代码
                                    class_obj = Class(
                                        name=class_name,
                                        code=class_code,
                                        created_by=current_user.id
                                    )
                                    db.session.add(class_obj)
                                    print(f"创建新班级: {class_name} (代码: {class_code})")
                                
                                # 加入缓存
                                created_classes[class_name] = class_obj
                        
                        # 先刷新以获取ID
                        db.session.flush()
                        
                        # 班级关联操作在flush之后进行
                        if class_obj:
                            if role == 'student':
                                # 检查是否已经在班级中（学生用classes关系）
                                if class_obj not in user.classes:
                                    # 数据一致性检查：确保学生不会被错误地加入教师关系
                                    if class_obj in user.teaching_classes:
                                        print(f"⚠️ 发现数据不一致：学生 {real_name} 已在教师关系中，先移除")
                                        user.teaching_classes.remove(class_obj)
                                    
                                    user.classes.append(class_obj)
                                    print(f"将学生 {real_name} 分配到班级 {class_name}")
                            elif role == 'teacher':
                                # 检查是否已经在班级中（教师用teaching_classes关系）
                                if class_obj not in user.teaching_classes:
                                    # 数据一致性检查：确保教师不会被错误地加入学生关系
                                    if class_obj in user.classes:
                                        print(f"⚠️ 发现数据不一致：教师 {real_name} 已在学生关系中，先移除")
                                        user.classes.remove(class_obj)
                                    
                                    user.teaching_classes.append(class_obj)
                                    print(f"将教师 {real_name} 分配到班级 {class_name}")
                        
                        success_count += 1
                        print(f"成功处理用户: {real_name}")
                        
                    except Exception as e:
                        error_msg = f'第{row_num}行: {str(e)}'
                        errors.append(error_msg)
                        error_count += 1
                        print(f"处理第{row_num}行时出错: {e}")
                        import traceback
                        traceback.print_exc()
                        # 对于单个记录的错误，我们不回滚整个事务，只是跳过这个记录
                        continue
            
            # 只有在所有操作都成功时才提交
            if error_count == 0 or success_count > 0:
                db.session.commit()
                print(f"事务提交成功，成功导入 {success_count} 个用户")
                
                # 进行数据一致性检查
                validate_class_associations()
            else:
                db.session.rollback()
                print("由于存在错误，事务已回滚")
            
            return jsonify({
                'success': success_count > 0,
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors,
                'message': f'成功导入 {success_count} 个用户' + (f'，{error_count} 个错误' if error_count > 0 else '')
            })
            
        except Exception as e:
            db.session.rollback()
            error_msg = f'批量导入过程中发生错误: {str(e)}'
            print(error_msg)
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': error_msg})
        
    except Exception as e:
        print(f"文件处理错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'文件处理失败: {str(e)}'})

# 班级管理路由
@app.route('/admin/classes')
@login_required
@require_teacher_or_admin
def manage_classes():
    if current_user.is_super_admin:
        classes = Class.query.order_by(Class.created_at.desc()).all()
    else:
        # 教师只能看到自己的班级
        classes = current_user.teaching_classes
    
    return render_template('manage_classes.html', classes=classes)

@app.route('/admin/classes/add', methods=['GET', 'POST'])
@login_required
@require_role(UserRole.SUPER_ADMIN)
def add_class():
    if request.method == 'POST':
        name = request.form['name']
        code = request.form['code']
        description = request.form.get('description', '')
        grade = request.form.get('grade', '')
        
        # 检查班级代码是否已存在
        if Class.query.filter_by(code=code).first():
            flash('班级代码已存在')
            return render_template('add_class.html')
        
        # 创建新班级
        new_class = Class(
            name=name,
            code=code,
            description=description,
            grade=grade,
            created_by=current_user.id
        )
        
        db.session.add(new_class)
        db.session.commit()
        
        flash(f'班级 {name} 创建成功！现在可以在创建用户时将教师和学生分配到该班级')
        return redirect(url_for('manage_classes'))
    
    return render_template('add_class.html')

@app.route('/admin/classes/<int:class_id>/edit', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def edit_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限编辑此班级')
            return redirect(url_for('manage_classes'))
    
    if request.method == 'POST':
        class_obj.name = request.form['name']
        class_obj.code = request.form['code']
        class_obj.description = request.form.get('description', '')
        class_obj.grade = request.form.get('grade', '')
        class_obj.is_active = 'is_active' in request.form
        
        # 更新教师分配（只有超级管理员可以修改）
        if current_user.is_super_admin:
            teacher_ids = request.form.getlist('teachers')
            class_obj.teachers.clear()
            if teacher_ids:
                teachers = User.query.filter(User.id.in_(teacher_ids), User.role == UserRole.TEACHER).all()
                class_obj.teachers.extend(teachers)
        
        db.session.commit()
        flash(f'班级 {class_obj.name} 信息已更新')
        return redirect(url_for('manage_classes'))
    
    teachers = User.query.filter_by(role=UserRole.TEACHER).all()
    return render_template('edit_class.html', class_obj=class_obj, teachers=teachers)

@app.route('/admin/classes/<int:class_id>/delete', methods=['POST'])
@login_required
@require_teacher_or_admin
def delete_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        # 教师只能删除自己负责的空班级
        if current_user not in class_obj.teachers:
            flash('您没有权限删除此班级')
            return redirect(url_for('manage_classes'))
        
        # 检查班级是否为空（没有学生和作业）
        if class_obj.students or class_obj.assignments:
            flash('只能删除没有学生和作业的空班级')
            return redirect(url_for('manage_classes'))
    else:
        # 超级管理员删除班级时的清理工作
        # 移除所有学生关联
        class_obj.students.clear()
        # 移除所有教师关联
        class_obj.teachers.clear()
        
        # 删除相关作业和提交文件
        import os
        for assignment in class_obj.assignments:
            for submission in assignment.submissions:
                try:
                    if os.path.exists(submission.file_path):
                        os.remove(submission.file_path)
                except Exception as e:
                    print(f"删除文件失败: {e}")
    
    class_name = class_obj.name
    db.session.delete(class_obj)
    db.session.commit()
    
    flash(f'班级 "{class_name}" 已成功删除')
    return redirect(url_for('manage_classes'))

@app.route('/admin/classes/<int:class_id>/students')
@login_required
@require_teacher_or_admin
def class_students(class_id):
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('manage_classes'))
    
    # 获取可以添加到班级的学生（排除已在班级中的学生）
    current_student_ids = [student.id for student in class_obj.students]
    available_students = User.query.filter(
        User.role == UserRole.STUDENT,
        ~User.id.in_(current_student_ids) if current_student_ids else True
    ).order_by(User.real_name).all()
    
    return render_template('class_students.html', 
                         class_obj=class_obj, 
                         available_students=available_students)

@app.route('/admin/classes/<int:class_id>/add_student', methods=['POST'])
@login_required
@require_teacher_or_admin
def add_student_to_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    student_id = request.form.get('student_id')
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('manage_classes'))
    
    student = User.query.filter_by(id=student_id, role=UserRole.STUDENT).first()
    if not student:
        flash('学生不存在')
        return redirect(url_for('class_students', class_id=class_id))
    
    if student in class_obj.students:
        flash(f'{student.real_name} 已经在班级中')
    else:
        class_obj.students.append(student)
        db.session.commit()
        flash(f'已将 {student.real_name} 添加到班级')
    
    return redirect(url_for('class_students', class_id=class_id))

@app.route('/admin/classes/<int:class_id>/remove_student', methods=['POST'])
@login_required
@require_teacher_or_admin
def remove_student_from_class(class_id):
    class_obj = Class.query.get_or_404(class_id)
    student_id = request.form.get('student_id')
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('manage_classes'))
    
    if student in class_obj.students:
        class_obj.students.remove(student)
        db.session.commit()
        flash(f'已将 {student.real_name} 从班级中移除')
    else:
        flash(f'{student.real_name} 不在此班级中')
    
    return redirect(url_for('class_students', class_id=class_id))

@app.route('/admin/classes/<int:class_id>/grades')
@login_required
@require_teacher_or_admin
def class_grades(class_id):
    """班级成绩统计页面"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限查看此班级成绩')
            return redirect(url_for('manage_classes'))
    
    # 获取班级所有作业
    assignments = Assignment.query.filter_by(class_id=class_id).order_by(Assignment.created_at).all()
    
    # 获取班级所有学生
    students = class_obj.students
    
    # 构建成绩统计数据
    grade_stats = []
    for student in students:
        student_data = {
            'student': student,
            'grades': {},
            'graded_count': 0,
            'average': 0
        }
        
        total_average = 0
        graded_assignments = 0
        
        for assignment in assignments:
            # 使用新的评分系统获取平均分
            average_grade = get_student_assignment_average_grade(assignment.id, student.id)
            
            if average_grade is not None:
                student_data['grades'][assignment.id] = average_grade
                total_average += average_grade
                graded_assignments += 1
            else:
                # 如果新系统没有评分，尝试从旧系统获取
                submission = Submission.query.filter_by(
                    assignment_id=assignment.id,
                    student_id=student.id
                ).filter(Submission.grade.isnot(None)).order_by(
                    Submission.graded_at.desc()
                ).first()
                
                if submission and submission.grade is not None:
                    student_data['grades'][assignment.id] = submission.grade
                    total_average += submission.grade
                    graded_assignments += 1
        
        # 计算平均分
        if graded_assignments > 0:
            student_data['average'] = round(total_average / graded_assignments, 2)
            student_data['graded_count'] = graded_assignments
        
        grade_stats.append(student_data)
    
    # 按平均分排序（而不是总分）
    grade_stats.sort(key=lambda x: x['average'], reverse=True)
    
    # 添加排名
    for rank, student_data in enumerate(grade_stats, 1):
        student_data['rank'] = rank
    
    return render_template('class_grades.html', 
                         class_obj=class_obj, 
                         assignments=assignments,
                         grade_stats=grade_stats)

@app.route('/admin/classes/<int:class_id>/export_grades')
@login_required
@require_teacher_or_admin
def export_class_grades(class_id):
    """导出班级成绩为Excel文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from flask import send_file
    
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限导出此班级成绩')
            return redirect(url_for('manage_classes'))
    
    # 获取数据（复用class_grades的逻辑）
    assignments = Assignment.query.filter_by(class_id=class_id).order_by(Assignment.created_at).all()
    students = class_obj.students
    
    grade_stats = []
    for student in students:
        student_data = {
            'student': student,
            'grades': {},
            'graded_count': 0,
            'average': 0
        }
        
        total_average = 0
        graded_assignments = 0
        
        for assignment in assignments:
            # 使用新的评分系统获取平均分
            average_grade = get_student_assignment_average_grade(assignment.id, student.id)
            
            if average_grade is not None:
                student_data['grades'][assignment.id] = average_grade
                total_average += average_grade
                graded_assignments += 1
            else:
                # 如果新系统没有评分，尝试从旧系统获取
                submission = Submission.query.filter_by(
                    assignment_id=assignment.id,
                    student_id=student.id
                ).filter(Submission.grade.isnot(None)).order_by(
                    Submission.graded_at.desc()
                ).first()
                
                if submission and submission.grade is not None:
                    student_data['grades'][assignment.id] = submission.grade
                    total_average += submission.grade
                    graded_assignments += 1
        
        if graded_assignments > 0:
            student_data['average'] = round(total_average / graded_assignments, 2)
            student_data['graded_count'] = graded_assignments
        
        grade_stats.append(student_data)
    
    # 按平均分排序（而不是总分）
    grade_stats.sort(key=lambda x: x['average'], reverse=True)
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name}_成绩单"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入标题
    ws['A1'] = f"{class_obj.name} 成绩统计表"
    ws.merge_cells('A1:' + chr(ord('E') + len(assignments)) + '1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_alignment
    
    # 写入统计信息
    beijing_now = datetime.now(BEIJING_TZ)
    ws['A2'] = f"导出时间：{beijing_now.strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"班级代码：{class_obj.code}"
    ws['A4'] = f"学生人数：{len(students)}"
    ws['A5'] = f"作业数量：{len(assignments)}"
    
    # 写入表头
    headers = ['排名', '姓名', '学号']
    for assignment in assignments:
        headers.append(assignment.title[:10] + ('...' if len(assignment.title) > 10 else ''))
    headers.append('平均分')  # 只保留平均分，移除总分
    
    row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # 写入数据
    for rank, student_data in enumerate(grade_stats, 1):
        row += 1
        student = student_data['student']
        
        # 基本信息
        data_row = [
            rank,
            student.real_name,
            student.student_id or student.username
        ]
        
        # 各作业成绩
        for assignment in assignments:
            grade = student_data['grades'].get(assignment.id, '')
            data_row.append(grade if grade != '' else '未评分')
        
        # 只添加平均分，移除总分
        data_row.append(student_data['average'] if student_data['graded_count'] > 0 else '')
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = center_alignment
            cell.border = border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8  # 排名
    ws.column_dimensions['B'].width = 15  # 姓名
    ws.column_dimensions['C'].width = 15  # 学号
    
    for i, assignment in enumerate(assignments, 4):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = 12  # 作业列
    
    ws.column_dimensions[chr(ord('A') + len(headers) - 2)].width = 10  # 总分
    ws.column_dimensions[chr(ord('A') + len(headers) - 1)].width = 10  # 平均分
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{class_obj.name}_成绩单_{beijing_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/assignment/create', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def create_assignment():
    if request.method == 'POST':
        print("DEBUG: create_assignment POST request received")
        title = request.form['title']
        description = request.form['description']
        due_date_str = request.form['due_date']
        allowed_file_types = request.form.get('allowed_file_types', '')
        max_file_size = request.form.get('max_file_size', '50')
        max_submissions = request.form.get('max_submissions', '0')  # 默认无限制提交
        class_id = request.form.get('class_id')  # 新增班级选择
        
        # 处理附件上传
        attachment_filename = None
        attachment_original_filename = None
        attachment_file_path = None
        attachment_file_size = None
        
        print(f"DEBUG: Checking for attachment in request.files: {list(request.files.keys())}")
        if 'attachment' in request.files:
            attachment_file = request.files['attachment']
            print(f"DEBUG: Attachment file found: {attachment_file.filename}")
            if attachment_file and attachment_file.filename:
                print("DEBUG: Saving attachment file")
                attachment_filename, attachment_original_filename, attachment_file_path, attachment_file_size = save_assignment_attachment(attachment_file)
                print(f"DEBUG: Attachment saved - filename: {attachment_filename}, original: {attachment_original_filename}, path: {attachment_file_path}, size: {attachment_file_size}")
            else:
                print("DEBUG: Attachment file is empty")
        else:
            print("DEBUG: No attachment in request.files")
        
        # 验证班级权限
        if class_id:
            selected_class = Class.query.get(class_id)
            if not selected_class:
                flash('选择的班级不存在')
                return render_template('create_assignment.html', available_classes=get_available_classes())
            
            # 权限检查：超级管理员可以为任何班级创建作业，教师只能为自己负责的班级创建
            if not current_user.is_super_admin and current_user not in selected_class.teachers:
                flash('您没有权限为此班级创建作业')
                return render_template('create_assignment.html', available_classes=get_available_classes())
        
        # 处理截止时间
        due_date = None
        if due_date_str:
            try:
                # 前端传来的是北京时间，需要转换为UTC时间存储
                local_time = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M')
                # 将北京时间转换为UTC时间（减去8小时）
                due_date = local_time - timedelta(hours=8)
            except ValueError:
                flash('日期格式错误')
                return render_template('create_assignment.html', available_classes=get_available_classes())
        
        # 处理文件大小限制（MB转换为字节）
        try:
            max_size_mb = float(max_file_size)
            # 验证文件大小在合理范围内（1MB到10GB）
            if max_size_mb < 1:
                flash('文件大小不能小于1MB')
                return render_template('create_assignment.html', available_classes=get_available_classes())
            elif max_size_mb > 10240:  # 10GB = 10240MB
                flash('文件大小不能超过10GB (10240MB)')
                return render_template('create_assignment.html', available_classes=get_available_classes())
            max_size_bytes = int(max_size_mb * 1024 * 1024)
        except (ValueError, TypeError):
            max_size_bytes = 50 * 1024 * 1024  # 默认50MB
        
        # 处理提交次数限制
        try:
            max_submissions_count = int(max_submissions)
            if max_submissions_count < 0:
                max_submissions_count = 0  # 0表示无限制
        except (ValueError, TypeError):
            max_submissions_count = 0  # 默认无限制提交
        
        # 处理文件类型
        if allowed_file_types:
            # 清理文件类型字符串，移除空格和点号
            file_types = []
            for ext in allowed_file_types.split(','):
                ext = ext.strip().lower()
                if ext.startswith('.'):
                    ext = ext[1:]
                if ext:
                    file_types.append(ext)
            allowed_file_types = ','.join(file_types)
        
        assignment = Assignment(
            title=title,
            description=description,
            due_date=due_date,
            allowed_file_types=allowed_file_types,
            max_file_size=max_size_bytes,
            max_submissions=max_submissions_count,  # 新增提交次数限制
            teacher_id=current_user.id,
            class_id=class_id if class_id else None,  # 绑定班级
            attachment_filename=attachment_filename,
            attachment_original_filename=attachment_original_filename,
            attachment_file_path=attachment_file_path,
            attachment_file_size=attachment_file_size
        )
        
        db.session.add(assignment)
        db.session.commit()
        flash('作业创建成功')
        
        # 根据用户角色重定向
        if current_user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        else:
            return redirect(url_for('teacher_dashboard'))
    
    # GET请求：获取可用的班级列表
    available_classes = get_available_classes()
    return render_template('create_assignment.html', available_classes=available_classes)

def get_available_classes():
    """获取当前用户可用的班级列表"""
    if current_user.is_super_admin:
        # 超级管理员可以看到所有活跃班级
        return Class.query.filter_by(is_active=True).all()
    elif current_user.is_teacher:
        # 教师只能看到自己负责的班级
        return current_user.teaching_classes
    else:
        return []

@app.route('/admin/assignment/<int:assignment_id>/submissions')
@login_required
@require_teacher_or_admin
def view_submissions(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查：超级管理员可以查看所有作业，教师可以查看自己管理的作业
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限查看此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 获取所有提交记录并按学生分组
    submissions = Submission.query.filter_by(assignment_id=assignment_id).order_by(Submission.submitted_at.desc()).all()
    
    # 按学生分组统计，但只处理 student_id 不为 None 的提交记录
    student_submissions = defaultdict(list)
    for submission in submissions:
        # 跳过 student_id 为 None 的提交记录
        if submission.student_id is None:
            continue
        student_key = (submission.student_id, submission.student_name, submission.student_number)
        student_submissions[student_key].append(submission)
    
    # 构建学生提交统计数据
    student_stats = []
    for (student_id, student_name, student_number), student_subs in student_submissions.items():
        # 获取最新提交和评分信息
        latest_submission = student_subs[0]  # 已按时间降序排列
        submission_count = len(student_subs)
        
        # 获取最新评分（使用新的评分系统）
        latest_grade = get_student_assignment_average_grade(assignment_id, student_id)
        latest_feedback = None
        
        # 获取最近的反馈（从任意教师）
        latest_grade_record = AssignmentGrade.query.filter(
            AssignmentGrade.assignment_id == assignment_id,
            AssignmentGrade.student_id == student_id,
            AssignmentGrade.feedback.isnot(None),
            AssignmentGrade.feedback != ''
        ).order_by(AssignmentGrade.updated_at.desc()).first()
        
        if latest_grade_record:
            latest_feedback = latest_grade_record.feedback
        
        # 如果新系统没有评分，则尝试从旧系统获取
        if latest_grade is None:
            graded_submissions = [s for s in student_subs if s.grade is not None]
            if graded_submissions:
                latest_graded = graded_submissions[0]
                latest_grade = latest_graded.grade
                if not latest_feedback:
                    latest_feedback = latest_graded.feedback
        
        student_stats.append({
            'student_id': student_id,
            'student_name': student_name,
            'student_number': student_number,
            'submission_count': submission_count,
            'latest_submission': latest_submission,
            'latest_grade': latest_grade,
            'latest_feedback': latest_feedback,
            'all_submissions': student_subs
        })
    
    # 按学生姓名排序
    student_stats.sort(key=lambda x: x['student_name'])
    
    return render_template('submissions.html', assignment=assignment, student_stats=student_stats)

@app.route('/student/assignment/<int:assignment_id>/submissions')
@login_required
@require_role(UserRole.STUDENT)
def student_submission_history(assignment_id):
    """学生查看自己对某个作业的提交记录（无截止时间限制）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 检查学生是否有权限查看此作业
    if assignment.class_id:
        # 如果作业绑定了班级，检查学生是否在该班级中
        class_obj = Class.query.get(assignment.class_id)
        if current_user not in class_obj.students:
            flash('您没有权限查看此作业')
            return redirect(url_for('student_dashboard'))
    
    # 获取学生的所有提交记录，按提交时间降序排列
    submission_history = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=current_user.id
    ).order_by(Submission.submitted_at.desc()).all()
    
    # 为每个提交记录添加评分教师信息
    for submission in submission_history:
        if submission.graded_by:
            submission.grader = User.query.get(submission.graded_by)
        else:
            submission.grader = None
    
    # 获取教师评分情况（新的评分系统）
    teacher_grades = get_student_assignment_teacher_grades(assignment_id, current_user.id)
    
    # 计算平均分
    average_grade = get_student_assignment_average_grade(assignment_id, current_user.id)
    
    return render_template('student_submission_history.html', 
                         assignment=assignment, 
                         submission_history=submission_history,
                         teacher_grades=teacher_grades,
                         average_grade=average_grade)

@app.route('/student/assignment/<int:assignment_id>/student/<int:student_id>/submissions')
@login_required
def student_submission_history_with_student_id(assignment_id, student_id):
    """通过学生 ID 查看提交记录（供教师使用）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if current_user.is_student:
        # 学生只能查看自己的记录
        if current_user.id != student_id:
            flash('您没有权限查看其他学生的提交记录')
            return redirect(url_for('student_dashboard'))
    elif not (current_user.is_super_admin or current_user.is_teacher):
        flash('您没有权限查看此内容')
        return redirect(url_for('index'))
    
    # 获取学生的所有提交记录
    submission_history = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).order_by(Submission.submitted_at.desc()).all()
    
    # 为每个提交记录添加评分教师信息
    for submission in submission_history:
        if submission.graded_by:
            submission.grader = User.query.get(submission.graded_by)
        else:
            submission.grader = None
    
    # 获取教师评分情况（新的评分系统）
    teacher_grades = get_student_assignment_teacher_grades(assignment_id, student_id)
    
    # 计算平均分
    average_grade = get_student_assignment_average_grade(assignment_id, student_id)
    
    return render_template('student_submission_history.html', 
                         assignment=assignment, 
                         submission_history=submission_history,
                         viewed_student=student,
                         teacher_grades=teacher_grades,
                         average_grade=average_grade)

@app.route('/admin/assignment/<int:assignment_id>/student/<int:student_id>/grade_assignment', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def grade_assignment_overall(assignment_id, student_id):
    """教师给学生的整个作业进行评分（新的评分机制）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限评分此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 获取该学生的所有提交记录
    submissions = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).order_by(Submission.submitted_at.desc()).all()
    
    if not submissions:
        flash('该学生尚未提交此作业')
        return redirect(url_for('view_submissions', assignment_id=assignment_id))
    
    # 获取当前教师对此作业的评分记录
    existing_grade = AssignmentGrade.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id,
        teacher_id=current_user.id
    ).first()
    
    if request.method == 'POST':
        grade = request.form.get('grade')
        feedback = request.form.get('feedback', '')
        
        # 验证评分
        grade_float = None
        if grade:
            try:
                grade_float = float(grade)
                if grade_float < 0 or grade_float > 100:
                    flash('评分必须在0-100之间')
                    return render_template('grade_assignment_overall.html', 
                                         assignment=assignment, 
                                         student=student, 
                                         submissions=submissions,
                                         existing_grade=existing_grade)
            except ValueError:
                flash('评分必须是有效的数字')
                return render_template('grade_assignment_overall.html', 
                                     assignment=assignment, 
                                     student=student, 
                                     submissions=submissions,
                                     existing_grade=existing_grade)
        
        # 创建或更新评分记录
        if existing_grade:
            existing_grade.grade = grade_float
            existing_grade.feedback = feedback
            existing_grade.updated_at = datetime.utcnow()
        else:
            existing_grade = AssignmentGrade(
                assignment_id=assignment_id,
                student_id=student_id,
                teacher_id=current_user.id,
                grade=grade_float,
                feedback=feedback
            )
            db.session.add(existing_grade)
        
        db.session.commit()
        flash(f'已成功给 {student.real_name} 的作业评分')
        
        return redirect(url_for('view_submissions', assignment_id=assignment_id))
    
    return render_template('grade_assignment_overall.html', 
                         assignment=assignment, 
                         student=student, 
                         submissions=submissions,
                         existing_grade=existing_grade)

@app.route('/admin/assignment/<int:assignment_id>/student/<int:student_id>/grade', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def grade_student_submissions(assignment_id, student_id):
    """教师给学生的作业进行评分"""
    assignment = Assignment.query.get_or_404(assignment_id)
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限评分此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 获取该学生的所有提交记录
    submissions = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).order_by(Submission.submitted_at.desc()).all()
    
    if not submissions:
        flash('该学生尚未提交此作业')
        return redirect(url_for('view_submissions', assignment_id=assignment_id))
    
    if request.method == 'POST':
        submission_id = request.form.get('submission_id')
        grade = request.form.get('grade')
        feedback = request.form.get('feedback', '')
        
        if not submission_id:
            flash('请选择要评分的提交记录')
            return render_template('grade_student_submissions.html', 
                                 assignment=assignment, 
                                 student=student, 
                                 submissions=submissions)
        
        submission = Submission.query.get_or_404(submission_id)
        
        # 验证评分
        if grade:
            try:
                grade_float = float(grade)
                if grade_float < 0 or grade_float > 100:
                    flash('评分必须在0-100之间')
                    return render_template('grade_student_submissions.html', 
                                         assignment=assignment, 
                                         student=student, 
                                         submissions=submissions)
                submission.grade = grade_float
            except ValueError:
                flash('评分必须是有效的数字')
                return render_template('grade_student_submissions.html', 
                                     assignment=assignment, 
                                     student=student, 
                                     submissions=submissions)
        else:
            submission.grade = None
        
        submission.feedback = feedback
        submission.graded_by = current_user.id
        submission.graded_at = datetime.utcnow()
        
        db.session.commit()
        flash(f'已成功评分 {student.real_name} 的作业')
        
        return redirect(url_for('view_submissions', assignment_id=assignment_id))
    
    return render_template('grade_student_submissions.html', 
                         assignment=assignment, 
                         student=student, 
                         submissions=submissions)

def get_student_assignment_average_grade(assignment_id, student_id):
    """获取学生作业的平均分（所有评分教师包括超级管理员的平均分）"""
    assignment = Assignment.query.get(assignment_id)
    if not assignment:
        return None
    
    # 获取该作业的所有评分教师（包括超级管理员）
    teacher_ids = []
    
    if assignment.class_id:
        class_obj = Class.query.get(assignment.class_id)
        if class_obj:
            # 班级的所有授课教师
            teacher_ids.extend([t.id for t in class_obj.teachers])
    
    # 添加作业创建者（如果不在授课教师列表中）
    if assignment.teacher_id not in teacher_ids:
        teacher_ids.append(assignment.teacher_id)
    
    # 添加所有超级管理员
    super_admins = User.query.filter_by(role=UserRole.SUPER_ADMIN).all()
    for admin in super_admins:
        if admin.id not in teacher_ids:
            teacher_ids.append(admin.id)
    
    # 获取所有评分教师（包括超级管理员）的评分
    grades = AssignmentGrade.query.filter(
        AssignmentGrade.assignment_id == assignment_id,
        AssignmentGrade.student_id == student_id,
        AssignmentGrade.teacher_id.in_(teacher_ids),
        AssignmentGrade.grade.isnot(None)
    ).all()
    
    if not grades:
        return None
    
    # 计算平均分：所有评分教师的分数总和除以教师数量
    total_grade = sum(grade.grade for grade in grades)
    return round(total_grade / len(grades), 2)

def get_student_assignment_teacher_grades(assignment_id, student_id):
    """获取学生作业的所有教师评分记录"""
    return AssignmentGrade.query.filter(
        AssignmentGrade.assignment_id == assignment_id,
        AssignmentGrade.student_id == student_id
    ).join(User, AssignmentGrade.teacher_id == User.id).all()

def can_teacher_manage_student(teacher, student):
    """检查教师是否有权限管理某个学生"""
    # 只能管理学生角色
    if student.role != UserRole.STUDENT:
        return False
    
    # 超级管理员可以管理所有学生
    if teacher.is_super_admin:
        return True
    
    # 教师只能管理自己班级中的学生
    if teacher.is_teacher:
        # 获取教师负责的所有班级
        teacher_classes = teacher.teaching_classes
        
        # 检查学生是否在这些班级中
        for class_obj in teacher_classes:
            if student in class_obj.students:
                return True
    
    return False

def can_manage_assignment(user, assignment):
    """检查用户是否可以管理此作业"""
    # 超级管理员可以管理所有作业
    if user.is_super_admin:
        return True
    
    # 教师可以管理自己创建的作业
    if assignment.teacher_id == user.id:
        return True
    
    # 教师可以管理分配给自己负责班级的作业
    if user.is_teacher and assignment.class_id:
        teacher_class_ids = [c.id for c in user.teaching_classes]
        if assignment.class_id in teacher_class_ids:
            return True
    
    return False

@app.route('/admin/assignment/<int:assignment_id>/edit', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def edit_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限编辑此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    if request.method == 'POST':
        # 获取表单数据
        title = request.form['title']
        description = request.form.get('description', '')
        due_date_str = request.form.get('due_date')
        file_types = request.form.get('file_types', '')
        max_size = request.form.get('max_size', '50')
        max_submissions = request.form.get('max_submissions', '0')  # 默认无限制提交
        class_id = request.form.get('class_id')  # 新增班级选择
        
        # 处理附件上传
        if 'attachment' in request.files:
            attachment_file = request.files['attachment']
            if attachment_file and attachment_file.filename:
                # 删除旧附件
                if assignment.attachment_file_path:
                    delete_assignment_attachment(assignment.attachment_file_path)
                
                # 保存新附件
                attachment_filename, attachment_original_filename, attachment_file_path, attachment_file_size = save_assignment_attachment(attachment_file)
                
                # 更新附件信息
                assignment.attachment_filename = attachment_filename
                assignment.attachment_original_filename = attachment_original_filename
                assignment.attachment_file_path = attachment_file_path
                assignment.attachment_file_size = attachment_file_size
        
        # 检查是否需要删除附件
        if 'delete_attachment' in request.form and request.form['delete_attachment'] == 'on':
            if assignment.attachment_file_path:
                delete_assignment_attachment(assignment.attachment_file_path)
                assignment.attachment_filename = None
                assignment.attachment_original_filename = None
                assignment.attachment_file_path = None
                assignment.attachment_file_size = None
                attachment_updated = True
        
        # 验证班级权限
        if class_id:
            selected_class = Class.query.get(class_id)
            if not selected_class:
                flash('选择的班级不存在')
                return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
            
            # 权限检查：超级管理员可以修改任何班级，教师只能修改自己负责的班级
            if not current_user.is_super_admin and current_user not in selected_class.teachers:
                flash('您没有权限将作业分配到此班级')
                return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
        
        # 验证和处理数据
        try:
            max_size_mb = float(max_size)
            # 验证文件大小在合理范围内（1MB到10GB）
            if max_size_mb < 1:
                flash('文件大小不能小于1MB')
                return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
            elif max_size_mb > 10240:  # 10GB = 10240MB
                flash('文件大小不能超过10GB (10240MB)')
                return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
            max_size_bytes = int(max_size_mb * 1024 * 1024)
        except (ValueError, TypeError):
            flash('文件大小限制必须是有效的数字')
            return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
        
        # 处理提交次数限制
        try:
            max_submissions_count = int(max_submissions)
            if max_submissions_count < 0:
                max_submissions_count = 0  # 0表示无限制
        except (ValueError, TypeError):
            max_submissions_count = 0  # 默认无限制提交
        
        # 处理截止时间
        due_date = None
        if due_date_str:
            try:
                # 前端传来的是北京时间，需要转换为UTC时间存储
                local_time = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M')
                # 将北京时间转换为UTC时间（减去8小时）
                due_date = local_time - timedelta(hours=8)
            except ValueError:
                flash('截止时间格式不正确')
                return render_template('edit_assignment.html', assignment=assignment, available_classes=get_available_classes())
        
        # 处理文件类型
        allowed_file_types = ''
        if file_types:
            file_types_list = []
            for ext in file_types.split(','):
                ext = ext.strip().lower()
                if ext.startswith('.'):
                    ext = ext[1:]
                if ext:
                    file_types_list.append(ext)
            allowed_file_types = ','.join(file_types_list)
        
        # 更新作业信息
        assignment.title = title
        assignment.description = description
        assignment.due_date = due_date
        assignment.allowed_file_types = allowed_file_types
        assignment.max_file_size = max_size_bytes
        assignment.max_submissions = max_submissions_count  # 更新提交次数限制
        assignment.class_id = class_id if class_id else None  # 更新班级绑定
        
        db.session.commit()
        flash('作业信息已成功更新')
        
        # 根据用户角色重定向
        if current_user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        else:
            return redirect(url_for('teacher_dashboard'))
    
    # GET请求：获取可用的班级列表
    available_classes = get_available_classes()
    return render_template('edit_assignment.html', assignment=assignment, available_classes=available_classes)

@app.route('/admin/assignment/<int:assignment_id>/delete', methods=['POST'])
@login_required
@require_teacher_or_admin
def delete_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限删除此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 删除相关的提交文件
    import os
    for submission in assignment.submissions:
        try:
            if os.path.exists(submission.file_path):
                os.remove(submission.file_path)
        except Exception as e:
            print(f"删除文件失败: {e}")
    
    # 删除作业附件
    if assignment.attachment_file_path:
        delete_assignment_attachment(assignment.attachment_file_path)
    
    assignment_title = assignment.title
    db.session.delete(assignment)
    db.session.commit()
    
    flash(f'作业 "{assignment_title}" 及其所有提交已成功删除')
    
    # 根据用户角色重定向
    if current_user.is_super_admin:
        return redirect(url_for('super_admin_dashboard'))
    else:
        return redirect(url_for('teacher_dashboard'))

@app.route('/admin/reset-system', methods=['GET', 'POST'])
@login_required
def reset_system():
    # 只有超级管理员可以重置系统
    if not current_user.is_super_admin:
        flash('您没有权限访问此功能')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        reset_type = request.form.get('reset_type')
        confirm_password = request.form.get('confirm_password')
        
        # 验证管理员密码
        if not current_user.check_password(confirm_password):
            flash('密码验证失败，重置操作已取消')
            return render_template('reset_system.html')
        
        try:
            if reset_type == 'assignments':
                # 清除作业数据
                import os
                import shutil
                
                try:
                    # 清空上传文件夹中的所有文件，但不删除文件夹本身
                    if os.path.exists(app.config['UPLOAD_FOLDER']):
                        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
                            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                            try:
                                if os.path.isfile(file_path) or os.path.islink(file_path):
                                    os.unlink(file_path)
                                elif os.path.isdir(file_path):
                                    shutil.rmtree(file_path)
                            except Exception as e:
                                app.logger.error(f'删除文件失败 {file_path}: {str(e)}')
                    
                    # 删除数据库中的相关记录（按正确顺序删除以避免外键约束问题）
                    # 1. 先删除作业评分记录
                    AssignmentGrade.query.delete()
                    db.session.commit()
                    
                    # 2. 再删除提交记录
                    Submission.query.delete()
                    db.session.commit()
                    
                    # 3. 最后删除作业
                    Assignment.query.delete()
                    db.session.commit()
                    
                    flash('作业数据已清除完成，所有作业、提交记录和评分记录已删除')
                except Exception as e:
                    db.session.rollback()
                    flash(f'清除作业数据失败: {str(e)}')
                    app.logger.error(f'清除作业数据失败: {str(e)}')
                
            elif reset_type == 'users':
                # 清除人员数据（保留超级管理员）
                admin_id = current_user.id
                
                # 删除非超级管理员用户相关的数据（按正确顺序删除以避免外键约束问题）
                
                # 1. 先删除作业评分记录（删除所有非超级管理员相关的评分）
                AssignmentGrade.query.filter(
                    AssignmentGrade.student_id != admin_id
                ).delete()
                AssignmentGrade.query.filter(
                    AssignmentGrade.teacher_id != admin_id
                ).delete()
                db.session.commit()
                
                # 2. 删除非超级管理员用户创建的提交记录
                submissions_to_delete = Submission.query.filter(
                    Submission.student_id != admin_id
                ).all()
                
                # 删除对应的文件
                import os
                for submission in submissions_to_delete:
                    try:
                        if os.path.exists(submission.file_path):
                            os.remove(submission.file_path)
                    except Exception as e:
                        app.logger.error(f"删除文件失败: {e}")
                
                # 删除提交记录
                for submission in submissions_to_delete:
                    db.session.delete(submission)
                db.session.commit()
                
                # 3. 删除非超级管理员的作业
                Assignment.query.filter(Assignment.teacher_id != admin_id).delete()
                db.session.commit()
                
                # 4. 删除班级（清除班级关联表）
                Class.query.delete()
                db.session.commit()
                
                # 5. 删除非超级管理员用户
                User.query.filter(
                    User.id != admin_id,
                    User.role != UserRole.SUPER_ADMIN
                ).delete()
                db.session.commit()
                
                flash('人员数据已清除完成，所有非超级管理员用户及其相关数据已删除')
                
            elif reset_type == 'all':
                # 重置所有数据
                import os
                import shutil
                
                # 清空上传文件夹中的所有文件，但不删除文件夹本身
                if os.path.exists(app.config['UPLOAD_FOLDER']):
                    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        try:
                            if os.path.isfile(file_path) or os.path.islink(file_path):
                                os.unlink(file_path)
                            elif os.path.isdir(file_path):
                                shutil.rmtree(file_path)
                        except Exception as e:
                            app.logger.error(f'删除文件失败 {file_path}: {str(e)}')
                
                # 保存当前超级管理员信息
                admin_username = current_user.username
                admin_real_name = current_user.real_name
                admin_password_hash = current_user.password_hash
                
                # 按正确顺序删除所有数据以避免外键约束问题
                # 1. 先删除作业评分记录
                AssignmentGrade.query.delete()
                db.session.commit()
                
                # 2. 再删除提交记录
                Submission.query.delete()
                db.session.commit()
                
                # 3. 删除作业
                Assignment.query.delete()
                db.session.commit()
                
                # 4. 删除班级（包括班级关联表class_student和class_teacher）
                Class.query.delete()
                db.session.commit()
                
                # 5. 最后删除用户
                User.query.delete()
                db.session.commit()
                
                # 重新创建超级管理员
                admin = User(
                    username=admin_username,
                    real_name=admin_real_name,
                    role=UserRole.SUPER_ADMIN
                )
                admin.password_hash = admin_password_hash
                db.session.add(admin)
                
                db.session.commit()
                
                # 重新登录
                logout_user()
                login_user(admin)
                
                flash('系统已完全重置，所有数据已清除并重新初始化')
            
            else:
                flash('无效的重置类型')
                
        except Exception as e:
            db.session.rollback()
            flash(f'重置操作失败: {str(e)}')
            app.logger.error(f'重置操作失败: {str(e)}')
        
        return redirect(url_for('super_admin_dashboard'))
    
    return render_template('reset_system.html')

@app.route('/submit/<int:assignment_id>', methods=['GET', 'POST'])
def submit_assignment(assignment_id):
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 检查作业是否已过截止时间
    if assignment.is_overdue():
        flash('很抱歉，该作业已过截止时间，无法提交')
        return redirect(url_for('index'))
    
    # 如果用户已登录，优先使用登录用户信息
    logged_in_student = None
    if current_user.is_authenticated and current_user.is_student:
        logged_in_student = current_user
        
        # 检查学生是否有权限提交此作业
        if assignment.class_id:  # 如果作业指定了班级
            student_classes = [c.id for c in current_user.classes]
            if assignment.class_id not in student_classes:
                flash('很抱歉，您不在此作业的指定班级中，无法提交')
                return redirect(url_for('student_dashboard'))
    
    if request.method == 'POST':
        # 再次检查作业是否已过截止时间（防止用户在提交过程中过期）
        if assignment.is_overdue():
            if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                return jsonify({'success': False, 'message': '很抱歉，该作业已过截止时间，无法提交'}), 400
            flash('很抱歉，该作业已过截止时间，无法提交')
            return redirect(url_for('student_dashboard') if logged_in_student else url_for('index'))
        
        # 优先使用登录用户信息
        if logged_in_student:
            student_name = logged_in_student.real_name
            student_number = logged_in_student.student_id or logged_in_student.username
            student_user_id = logged_in_student.id
        else:
            student_name = request.form['student_name']
            student_number = request.form['student_id']
            student_user_id = None
        
        # 检查是否还能提交
        if logged_in_student and not assignment.can_student_submit(student_user_id):
            flash(f'您已达到该作业的最大提交次数限制 ({assignment.max_submissions}次)')
            return redirect(url_for('student_dashboard'))
        
        notes = request.form.get('notes', '')
        
        if 'file' not in request.files:
            if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                return jsonify({'success': False, 'message': '请选择文件'}), 400
            flash('请选择文件')
            return render_template('submit.html', assignment=assignment, logged_in_student=logged_in_student)
        
        file = request.files['file']
        if file.filename == '':
            if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                return jsonify({'success': False, 'message': '请选择文件'}), 400
            flash('请选择文件')
            return render_template('submit.html', assignment=assignment, logged_in_student=logged_in_student)
        
        # 检查文件类型
        if not assignment.is_file_allowed(file.filename):
            allowed_types = ', '.join(assignment.get_allowed_extensions())
            error_msg = f'不允许的文件类型。允许的类型：{allowed_types}'
            if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg)
            return render_template('submit.html', assignment=assignment, logged_in_student=logged_in_student)
        
        # 检查文件大小
        file.seek(0, 2)  # 移动到文件末尾
        file_size = file.tell()
        file.seek(0)  # 重置文件指针
        
        if file_size > assignment.max_file_size:
            max_size_mb = assignment.max_file_size / (1024 * 1024)
            error_msg = f'文件大小超出限制。最大允许：{max_size_mb:.1f}MB'
            if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg)
            return render_template('submit.html', assignment=assignment, logged_in_student=logged_in_student)
        
        if file:
            try:
                # 生成安全的文件名 - 学生作业提交重命名格式：姓名-提交时间（年月日时分秒）-uuid
                original_filename = file.filename
                # 使用北京时间生成时间戳
                beijing_now = datetime.now(BEIJING_TZ)
                timestamp = beijing_now.strftime("%Y%m%d%H%M%S")
                filename_uuid = str(uuid.uuid4())[:8]  # 使用较短的UUID
                
                # 处理学生姓名，确保文件名安全
                safe_student_name = safe_chinese_filename(student_name)
                filename = f"{safe_student_name}-{timestamp}-{filename_uuid}{os.path.splitext(original_filename)[1]}"
                
                # 创建特定格式的文件夹 - 作业序号-作业名称-作业创建时间
                class_name = "无班级"  # 默认值
                if assignment.class_info:
                    class_name = assignment.class_info.name
                
                # 清理文件名中的非法字符（保留中文）
                safe_assignment_title = safe_chinese_filename(assignment.title)
                # 使用北京时间格式化日期
                assignment_beijing_time = to_beijing_time(assignment.created_at)
                assignment_date = assignment_beijing_time.strftime("%Y%m%d")
                
                # 作业序号-作业名称-作业创建时间
                folder_name = f"{assignment.id}-{safe_assignment_title}-{assignment_date}"
                folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
                os.makedirs(folder_path, exist_ok=True)
                
                # 保存文件到指定文件夹
                file_path = os.path.join(folder_path, filename)
                file.save(file_path)
                
                submission = Submission(
                    assignment_id=assignment_id,
                    student_id=student_user_id,
                    student_name=student_name,
                    student_number=student_number,
                    filename=filename,
                    original_filename=original_filename,
                    file_path=file_path,
                    file_size=file_size,
                    notes=notes
                )
                
                db.session.add(submission)
                db.session.commit()
                
                # 根据请求类型返回不同响应
                if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                    # Ajax请求，返回JSON
                    return jsonify({
                        'success': True, 
                        'message': '作业提交成功',
                        'redirect_url': url_for('student_dashboard') if logged_in_student else url_for('index')
                    })
                else:
                    # 普通表单提交
                    flash('作业提交成功')
                    
                    # 根据用户状态重定向
                    if logged_in_student:
                        return redirect(url_for('student_dashboard'))
                    else:
                        return redirect(url_for('index'))
                        
            except Exception as e:
                db.session.rollback()
                error_msg = f'文件上传失败: {str(e)}'
                if request.headers.get('Content-Type', '').startswith('multipart/form-data'):
                    return jsonify({'success': False, 'message': error_msg}), 500
                flash(error_msg)
                return render_template('submit.html', assignment=assignment, logged_in_student=logged_in_student)
    
    # GET请求：显示提交页面和历史记录
    submission_history = []
    if logged_in_student:
        submission_history = Submission.query.filter_by(
            assignment_id=assignment_id, 
            student_id=logged_in_student.id
        ).order_by(Submission.submitted_at.desc()).all()
        
        # 为每个提交记录添加评分教师信息
        for submission in submission_history:
            if submission.graded_by:
                submission.grader = User.query.get(submission.graded_by)
            else:
                submission.grader = None
    
    return render_template('submit.html', 
                         assignment=assignment, 
                         logged_in_student=logged_in_student,
                         submission_history=submission_history)

@app.route('/download/<int:submission_id>')
@login_required
@require_teacher_or_admin
def download_file(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    assignment = Assignment.query.get(submission.assignment_id)
    
    # 权限检查：使用统一的权限检查函数
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限下载此文件')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 检查文件是否存在
    if not os.path.exists(submission.file_path):
        flash('文件不存在或已被删除')
        return redirect(url_for('view_submissions', assignment_id=assignment.id))
    
    # 获取文件的目录和文件名
    file_directory = os.path.dirname(submission.file_path)
    filename = os.path.basename(submission.file_path)
    
    # 安全处理下载文件名，确保HTTP头兼容性
    try:
        # 尝试URL编码处理中文文件名
        from urllib.parse import quote
        safe_download_name = quote(submission.original_filename.encode('utf-8'))
        # 如果文件名过长或包含特殊字符，使用备用方案
        if len(safe_download_name) > 200:
            # 使用安全的文件名作为下载名
            file_ext = os.path.splitext(submission.original_filename)[1]
            safe_download_name = f"submission_{submission.id}{file_ext}"
        else:
            safe_download_name = submission.original_filename
    except:
        # 出现任何编码问题时，使用备用文件名
        file_ext = os.path.splitext(submission.original_filename)[1]
        safe_download_name = f"submission_{submission.id}{file_ext}"
    
    return send_from_directory(
        file_directory,
        filename,
        as_attachment=True,
        download_name=safe_download_name
    )

@app.route('/preview/<int:submission_id>')
@login_required
@require_teacher_or_admin
def preview_file(submission_id):
    """预览文件（主PDF）"""
    submission = Submission.query.get_or_404(submission_id)
    assignment = Assignment.query.get(submission.assignment_id)
    
    # 权限检查：使用统一的权限检查函数
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限预览此文件')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 检查文件是否存在
    if not os.path.exists(submission.file_path):
        flash('文件不存在或已被删除')
        return redirect(url_for('view_submissions', assignment_id=assignment.id))
    
    # 获取文件的目录和文件名
    file_directory = os.path.dirname(submission.file_path)
    filename = os.path.basename(submission.file_path)
    
    # 如果是PDF文件，返回适合浏览器预览的格式
    if submission.is_pdf():
        return send_from_directory(
            file_directory,
            filename,
            as_attachment=False,  # 不作为附件下载
            mimetype='application/pdf'
        )
    else:
        # 非PDF文件仍然作为下载
        return redirect(url_for('download_file', submission_id=submission_id))

@app.route('/api/assignment/<int:assignment_id>/info')
def get_assignment_info(assignment_id):
    """获取作业的最新信息（用于实时更新截止时间）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 统一时间格式，确保与前端完全一致
    due_date_utc = assignment.due_date.strftime('%Y-%m-%d %H:%M:%S') if assignment.due_date else None
    due_date_beijing = to_beijing_time(assignment.due_date).strftime('%Y-%m-%d %H:%M:%S') if assignment.due_date else None
    
    response_data = {
        'id': assignment.id,
        'title': assignment.title,
        'due_date': due_date_utc,  # UTC时间格式
        'due_date_beijing': due_date_beijing,  # 北京时间格式
        'is_overdue': assignment.is_overdue(),
        'last_updated': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')  # 添加最后更新时间
    }
    
    return jsonify(response_data)

@app.route('/admin/assignment/<int:assignment_id>/attachment')
@login_required
@require_teacher_or_admin
def download_assignment_attachment(assignment_id):
    """下载作业附件"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限下载此作业附件')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 检查是否有附件
    if not assignment.attachment_file_path or not os.path.exists(assignment.attachment_file_path):
        flash('附件不存在')
        return redirect(url_for('edit_assignment', assignment_id=assignment_id))
    
    # 发送文件
    return send_from_directory(
        directory=os.path.dirname(assignment.attachment_file_path),
        path=os.path.basename(assignment.attachment_file_path),
        as_attachment=True,
        download_name=assignment.attachment_original_filename or assignment.attachment_filename
    )

@app.route('/student/assignment/<int:assignment_id>/attachment')
@login_required
@require_role(UserRole.STUDENT)
def student_download_assignment_attachment(assignment_id):
    """学生下载作业附件"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 检查学生是否有权限查看此作业
    if assignment.class_id:
        # 如果作业绑定了班级，检查学生是否在该班级中
        class_obj = Class.query.get(assignment.class_id)
        if current_user not in class_obj.students:
            flash('您没有权限查看此作业')
            return redirect(url_for('student_dashboard'))
    
    # 检查是否有附件
    if not assignment.attachment_file_path or not os.path.exists(assignment.attachment_file_path):
        flash('附件不存在')
        return redirect(url_for('student_dashboard'))
    
    # 发送文件
    return send_from_directory(
        directory=os.path.dirname(assignment.attachment_file_path),
        path=os.path.basename(assignment.attachment_file_path),
        as_attachment=True,
        download_name=assignment.attachment_original_filename or assignment.attachment_filename
    )

@app.route('/admin/submission/<int:submission_id>/delete', methods=['POST'])
@login_required
def delete_submission(submission_id):
    # 只有超级管理员可以删除提交记录
    if not current_user.is_super_admin:
        flash('只有超级管理员才能删除提交记录')
        return redirect(url_for('index'))
    
    submission = Submission.query.get_or_404(submission_id)
    assignment = submission.assignment
    
    # 删除文件
    import os
    try:
        if os.path.exists(submission.file_path):
            os.remove(submission.file_path)
    except Exception as e:
        app.logger.error(f"删除文件失败: {e}")
    
    # 获取学生姓名用于显示消息
    student_name = submission.student_name
    submission_filename = submission.original_filename
    
    # 删除数据库记录
    db.session.delete(submission)
    db.session.commit()
    
    flash(f'已成功删除 {student_name} 的提交文件 "{submission_filename}"')
    
    # 重定向到作业提交查看页面
    return redirect(url_for('view_submissions', assignment_id=assignment.id))

@app.route('/admin/submission/<int:submission_id>/grade', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def grade_submission(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    assignment = submission.assignment
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限评分此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    if request.method == 'POST':
        # 获取评分和反馈
        grade = request.form.get('grade')
        feedback = request.form.get('feedback', '')
        
        # 验证评分
        if grade:
            try:
                grade_float = float(grade)
                if grade_float < 0 or grade_float > 100:
                    flash('评分必须在0-100之间')
                    return render_template('grade_submission.html', submission=submission, assignment=assignment)
                submission.grade = grade_float
            except ValueError:
                flash('评分必须是有效的数字')
                return render_template('grade_submission.html', submission=submission, assignment=assignment)
        else:
            submission.grade = None
        
        submission.feedback = feedback
        submission.graded_by = current_user.id
        submission.graded_at = datetime.utcnow()
        
        db.session.commit()
        flash('作业评分成功')
        
        # 重定向到作业提交列表
        return redirect(url_for('view_submissions', assignment_id=assignment.id))
    
    return render_template('grade_submission.html', submission=submission, assignment=assignment)

@app.route('/student/submission/<int:submission_id>/feedback')
@login_required
@require_role(UserRole.STUDENT)
def view_feedback(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    
    # 检查权限：只能查看自己的作业反馈
    if submission.student_id != current_user.id:
        flash('您没有权限查看此作业反馈')
        return redirect(url_for('student_dashboard'))
    
    return render_template('view_feedback.html', submission=submission)

@app.route('/admin/assignment/<int:assignment_id>/download_status')
@login_required
@require_teacher_or_admin
def download_assignment_status(assignment_id):
    """获取下载进度"""
    from flask import session
    progress_key = f'download_progress_{assignment_id}_{current_user.id}'
    progress = session.get(progress_key, {'status': 'pending', 'progress': 0, 'message': '准备中...'})
    return jsonify(progress)

@app.route('/admin/assignment/<int:assignment_id>/download')
@login_required
@require_teacher_or_admin
def download_assignment(assignment_id):
    """下载指定作业的所有提交文件"""
    import zipfile
    from io import BytesIO
    from flask import send_file, session
    import threading
    import time
    
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限下载此作业')
        return redirect(url_for('teacher_dashboard' if current_user.is_teacher else 'super_admin_dashboard'))
    
    # 获取作业的所有提交记录
    submissions = Submission.query.filter_by(assignment_id=assignment_id).all()
    
    if not submissions:
        flash('该作业没有任何提交记录')
        return redirect(url_for('view_submissions', assignment_id=assignment_id))
    
    # 进度跟踪键
    progress_key = f'download_progress_{assignment_id}_{current_user.id}'
    
    # 初始化进度
    session[progress_key] = {
        'status': 'processing',
        'progress': 0,
        'message': '正在检查文件...',
        'total_files': len(submissions)
    }
    session.permanent = True
    
    # 创建内存ZIP文件
    memory_file = BytesIO()
    
    try:
        # 使用最高压缩级别
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
            processed_files = 0
            total_files = len([s for s in submissions if os.path.exists(s.file_path)])
            
            for i, submission in enumerate(submissions):
                if os.path.exists(submission.file_path):
                    # 更新进度
                    progress_percent = int((processed_files / total_files) * 100) if total_files > 0 else 0
                    session[progress_key] = {
                        'status': 'processing',
                        'progress': progress_percent,
                        'message': f'正在压缩文件: {submission.original_filename}',
                        'current_file': processed_files + 1,
                        'total_files': total_files
                    }
                    
                    # 创建文件在ZIP中的路径：学生姓名_学号_提交时间_原文件名
                    beijing_time = to_beijing_time(submission.submitted_at)
                    time_str = beijing_time.strftime('%Y%m%d_%H%M%S') if beijing_time else 'unknown'
                    
                    safe_student_name = safe_chinese_filename(submission.student_name)
                    safe_original_name = safe_chinese_filename(submission.original_filename)
                    
                    zip_filename = f"{safe_student_name}_{submission.student_number}_{time_str}_{safe_original_name}"
                    
                    # 添加文件到ZIP
                    zf.write(submission.file_path, zip_filename)
                    processed_files += 1
                    
                    # 模拟小延迟，让进度条更可见（仅在文件少时）
                    if total_files < 10:
                        time.sleep(0.1)
        
        # 完成压缩
        session[progress_key] = {
            'status': 'completed',
            'progress': 100,
            'message': '压缩完成，准备下载...',
            'total_files': total_files
        }
        
        memory_file.seek(0)
        
        # 生成ZIP文件名：班级-作业标题-作业创建时间.zip
        beijing_created = to_beijing_time(assignment.created_at)
        created_time_str = beijing_created.strftime('%Y%m%d%H%M%S') if beijing_created else 'unknown'
        
        if assignment.class_info:
            class_name = safe_chinese_filename(assignment.class_info.name)
        else:
            class_name = '公共作业'
        
        safe_title = safe_chinese_filename(assignment.title)
        zip_filename = f"{class_name}-{safe_title}-{created_time_str}.zip"
        
        # 清理进度记录
        if progress_key in session:
            session.pop(progress_key)
        
        return send_file(
            memory_file,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        # 错误处理
        session[progress_key] = {
            'status': 'error',
            'progress': 0,
            'message': f'压缩失败: {str(e)}'
        }
        flash(f'下载失败: {str(e)}')
        return redirect(url_for('view_submissions', assignment_id=assignment_id))

@app.route('/admin/assignments/batch_download_status')
@login_required
@require_teacher_or_admin
def batch_download_status():
    """获取批量下载进度"""
    from flask import session
    import time
    
    progress_key = f'batch_download_progress_{current_user.id}'
    progress = session.get(progress_key, {'status': 'pending', 'progress': 0, 'message': '准备中...'})
    
    # 检查是否有启动时间，用于超时检测
    if 'start_time' in progress:
        elapsed_time = time.time() - progress['start_time']
        # 如果超过5分钟没有进度更新，认为下载已完成或失败
        if elapsed_time > 300 and progress.get('status') not in ['completed', 'error']:
            progress = {
                'status': 'completed',
                'progress': 100,
                'message': '下载已完成（检测到超时，可能已自动下载）',
                'timeout': True
            }
            # 超时时不立即清理进度记录，让前端可以检测到completed状态
    
    print(f"[调试] 批量下载进度查询: {progress}")  # 添加调试日志
    return jsonify(progress)

@app.route('/admin/assignments/batch_download_start', methods=['POST'])
@login_required
@require_teacher_or_admin
def start_batch_download():
    """启动批量下载进程"""
    from flask import session
    import time
    
    download_type = request.form.get('download_type')
    class_id = request.form.get('class_id')
    
    # 进度跟踪键
    progress_key = f'batch_download_progress_{current_user.id}'
    
    # 初始化进度
    session[progress_key] = {
        'status': 'started',
        'progress': 5,
        'message': '正在检查作业...',
        'download_type': download_type,
        'class_id': class_id,
        'start_time': time.time()  # 记录启动时间
    }
    session.permanent = True
    
    print(f"[调试] 初始化批量下载进度: {session[progress_key]}")  # 添加调试日志
    
    return jsonify({'success': True, 'message': '批量下载已启动'})

@app.route('/admin/assignments/batch_download_clear', methods=['POST'])
@login_required
@require_teacher_or_admin
def clear_batch_download_progress():
    """清理批量下载进度记录"""
    from flask import session
    
    progress_key = f'batch_download_progress_{current_user.id}'
    
    if progress_key in session:
        session.pop(progress_key)
        print(f"[调试] 已清理批量下载进度记录: {progress_key}")
    
    return jsonify({'success': True, 'message': '进度记录已清理'})

@app.route('/admin/assignments/batch_download', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def batch_download_assignments():
    """批量下载作业"""
    if request.method == 'GET':
        # 显示批量下载选择页面
        if current_user.is_super_admin:
            classes = Class.query.order_by(Class.name).all()
            assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
        else:
            # 教师只能看到自己的班级和作业
            classes = current_user.teaching_classes
            own_assignments = Assignment.query.filter_by(teacher_id=current_user.id).all()
            class_assignments = []
            if classes:
                class_ids = [c.id for c in classes]
                class_assignments = Assignment.query.filter(
                    Assignment.class_id.in_(class_ids),
                    Assignment.teacher_id != current_user.id
                ).all()
            assignments = own_assignments + class_assignments
        
        return render_template('batch_download.html', classes=classes, assignments=assignments)
    
    # POST请求：执行批量下载
    import zipfile
    from io import BytesIO
    from flask import send_file, session
    import time
    
    download_type = request.form.get('download_type')
    
    # 进度跟踪键
    progress_key = f'batch_download_progress_{current_user.id}'
    
    print(f"[调试] 批量下载开始处理: {download_type}")
    
    # 初始化进度
    session[progress_key] = {
        'status': 'processing',
        'progress': 10,
        'message': '正在检查作业...',
        'total_assignments': 0,
        'current_assignment': 0
    }
    session.permanent = True
    
    print(f"[调试] 批量下载开始处理: {download_type}")  # 添加调试日志
    
    if download_type == 'all':
        # 下载所有作业
        if current_user.is_super_admin:
            assignments = Assignment.query.all()
        else:
            # 教师只能下载自己的作业
            own_assignments = Assignment.query.filter_by(teacher_id=current_user.id).all()
            class_assignments = []
            teacher_classes = current_user.teaching_classes
            if teacher_classes:
                class_ids = [c.id for c in teacher_classes]
                class_assignments = Assignment.query.filter(
                    Assignment.class_id.in_(class_ids),
                    Assignment.teacher_id != current_user.id
                ).all()
            assignments = own_assignments + class_assignments
        
        zip_filename = f'所有作业-{datetime.now().strftime("%Y%m%d%H%M%S")}.zip'
        
    elif download_type == 'class':
        # 下载指定班级的所有作业
        class_id = request.form.get('class_id')
        if not class_id:
            flash('请选择班级')
            return redirect(url_for('batch_download_assignments'))
        
        class_obj = Class.query.get_or_404(class_id)
        
        # 权限检查
        if not current_user.is_super_admin and current_user not in class_obj.teachers:
            flash('您没有权限下载此班级的作业')
            return redirect(url_for('batch_download_assignments'))
        
        assignments = Assignment.query.filter_by(class_id=class_id).all()
        safe_class_name = safe_chinese_filename(class_obj.name)
        zip_filename = f'{safe_class_name}-所有作业-{datetime.now().strftime("%Y%m%d%H%M%S")}.zip'
        
    else:
        flash('无效的下载类型')
        return redirect(url_for('batch_download_assignments'))
    
    if not assignments:
        session[progress_key] = {
            'status': 'error',
            'progress': 0,
            'message': '没有找到任何作业'
        }
        flash('没有找到任何作业')
        return redirect(url_for('batch_download_assignments'))
    
    # 更新总作业数
    session[progress_key]['total_assignments'] = len(assignments)
    
    # 创建内存ZIP文件
    memory_file = BytesIO()
    
    try:
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
            for assignment_index, assignment in enumerate(assignments):
                # 更新当前作业进度
                assignment_progress = int((assignment_index / len(assignments)) * 100)
                session[progress_key] = {
                    'status': 'processing',
                    'progress': assignment_progress,
                    'message': f'正在处理作业: {assignment.title}',
                    'current_assignment': assignment_index + 1,
                    'total_assignments': len(assignments)
                }
                
                submissions = Submission.query.filter_by(assignment_id=assignment.id).all()
                
                if submissions:
                    # 为每个作业创建文件夹
                    beijing_created = to_beijing_time(assignment.created_at)
                    created_time_str = beijing_created.strftime('%Y%m%d%H%M%S') if beijing_created else 'unknown'
                    
                    if assignment.class_info:
                        class_name = safe_chinese_filename(assignment.class_info.name)
                    else:
                        class_name = '公共作业'
                    
                    safe_title = safe_chinese_filename(assignment.title)
                    assignment_folder = f"{class_name}-{safe_title}-{created_time_str}"
                    
                    for submission_index, submission in enumerate(submissions):
                        if os.path.exists(submission.file_path):
                            # 更新文件进度
                            file_progress = int(((assignment_index + (submission_index / len(submissions))) / len(assignments)) * 100)
                            session[progress_key] = {
                                'status': 'processing',
                                'progress': file_progress,
                                'message': f'正在压缩: {submission.original_filename}',
                                'current_assignment': assignment_index + 1,
                                'total_assignments': len(assignments),
                                'current_file': submission_index + 1,
                                'total_files': len(submissions)
                            }
                            
                            # 创建文件在ZIP中的路径
                            beijing_time = to_beijing_time(submission.submitted_at)
                            time_str = beijing_time.strftime('%Y%m%d_%H%M%S') if beijing_time else 'unknown'
                            
                            safe_student_name = safe_chinese_filename(submission.student_name)
                            safe_original_name = safe_chinese_filename(submission.original_filename)
                            
                            zip_filename_in_folder = f"{assignment_folder}/{safe_student_name}_{submission.student_number}_{time_str}_{safe_original_name}"
                            
                            # 添加文件到ZIP
                            zf.write(submission.file_path, zip_filename_in_folder)
                            
                            # 小延迟，让进度更可见
                            time.sleep(0.01)
        
        # 完成压缩
        session[progress_key] = {
            'status': 'completed',
            'progress': 100,
            'message': '所有作业压缩完成，开始下载...',
            'total_assignments': len(assignments)
        }
        
        memory_file.seek(0)
        
        # 注意：不在这里清理进度记录，让前端能检测到completed状态
        # 进度记录将在前端检测到completed状态后自动清理，或者通过超时机制清理
        
        return send_file(
            memory_file,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        # 错误处理
        session[progress_key] = {
            'status': 'error',
            'progress': 0,
            'message': f'批量下载失败: {str(e)}'
        }
        flash(f'批量下载失败: {str(e)}')
        return redirect(url_for('batch_download_assignments'))

def init_db():
    """初始化数据库和默认用户"""
    try:
        with app.app_context():
            db.create_all()
            
            # 创建默认超级管理员用户
            admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
            admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
            
            admin = User.query.filter_by(username=admin_username).first()
            if not admin:
                admin = User(
                    username=admin_username, 
                    real_name='超级管理员',
                    role=UserRole.SUPER_ADMIN,
                    must_change_password=False  # 超级管理员不需要强制修改密码
                )
                admin.set_password(admin_password)
                db.session.add(admin)
                db.session.commit()
                print(f"创建默认超级管理员用户: {admin_username}")
            else:
                # 更新旧用户的角色为超级管理员
                updated = False
                if not hasattr(admin, 'role') or admin.role != UserRole.SUPER_ADMIN:
                    admin.role = UserRole.SUPER_ADMIN
                    updated = True
                if not hasattr(admin, 'real_name') or not admin.real_name:
                    admin.real_name = '超级管理员'
                    updated = True
                if updated:
                    db.session.commit()
                print(f"超级管理员用户 {admin_username} 已存在")
    except Exception as e:
        print(f"数据库初始化错误: {e}")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)