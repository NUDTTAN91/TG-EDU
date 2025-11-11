"""评分相关路由"""
from datetime import datetime
from io import BytesIO
import pandas as pd
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, current_app
from flask_login import login_required, current_user

from app.extensions import db
from app.models import Assignment, Submission, User, Class, UserRole
from app.models.assignment import AssignmentGrade
from app.utils.decorators import require_teacher_or_admin
from app.services import NotificationService
from app.services.log_service import LogService
from app.utils import to_beijing_time

bp = Blueprint('grading', __name__, url_prefix='/admin')


def can_manage_assignment(user, assignment):
    """检查用户是否可以管理此作业"""
    # 超级管理员可以管理所有作业
    if user.is_super_admin:
        return True
    
    # 教师可以管理自己创建的作业
    if assignment.teacher_id == user.id:
        return True
    
    # 教师可以管理分配给自己负责班级的作业
    if user.is_teacher and assignment.class_id:
        teacher_class_ids = [c.id for c in user.teaching_classes]
        if assignment.class_id in teacher_class_ids:
            return True
    
    return False


@bp.route('/assignment/<int:assignment_id>/student/<int:student_id>/grade_assignment', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def grade_assignment_overall(assignment_id, student_id):
    """教师给学生的整个作业进行评分（新的评分机制）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    student = User.query.get_or_404(student_id)
    
    # 获取URL参数，判断是否从补交评分入口进入
    force_makeup = request.args.get('is_makeup', '0') == '1'
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限评分此作业')
        return redirect(url_for('admin.teacher_dashboard' if current_user.is_teacher else 'admin.super_admin_dashboard'))
    
    # 获取该学生的所有提交记录
    submissions = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).order_by(Submission.submitted_at.desc()).all()
    
    # 注意：即使学生没有提交，也允许老师打补交分
    # 不再强制要求必须有提交记录
    
    # 获取当前教师对此作业的评分记录
    existing_grade = AssignmentGrade.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id,
        teacher_id=current_user.id
    ).first()
    
    if request.method == 'POST':
        grade = request.form.get('grade')
        feedback = request.form.get('feedback', '')
        is_makeup = request.form.get('is_makeup') == 'on'  # 是否为补交
        is_cheating = request.form.get('is_cheating') == 'on'  # 是否作弊
        discount_rate = request.form.get('discount_rate', '100')  # 折扣百分比
        
        # 验证评分
        grade_float = None
        original_grade = None
        if grade:
            try:
                grade_float = float(grade)
                if grade_float < 0 or grade_float > 100:
                    flash('评分必须在0-100之间')
                    return render_template('grade_assignment_overall.html', 
                                         assignment=assignment, 
                                         student=student, 
                                         submissions=submissions,
                                         existing_grade=existing_grade,
                                         force_makeup=force_makeup)
                
                # 如果是补交，计算折扣后的分数
                if is_makeup:
                    try:
                        discount_rate_float = float(discount_rate)
                        if discount_rate_float < 0 or discount_rate_float > 100:
                            flash('折扣百分比必须在0-100之间')
                            return render_template('grade_assignment_overall.html', 
                                                 assignment=assignment, 
                                                 student=student, 
                                                 submissions=submissions,
                                                 existing_grade=existing_grade,
                                                 force_makeup=force_makeup)
                        original_grade = grade_float  # 保存原始分数
                        grade_float = round(grade_float * discount_rate_float / 100, 1)  # 计算折扣后分数
                        current_app.logger.info(f"[MAKEUP] 原始分数={original_grade}, 折扣={discount_rate_float}%, 最终分数={grade_float}")
                    except ValueError:
                        flash('折扣百分比必须是有效的数字')
                        return render_template('grade_assignment_overall.html', 
                                             assignment=assignment, 
                                             student=student, 
                                             submissions=submissions,
                                             existing_grade=existing_grade,
                                             force_makeup=force_makeup)
                    
            except ValueError:
                flash('评分必须是有效的数字')
                return render_template('grade_assignment_overall.html', 
                                     assignment=assignment, 
                                     student=student, 
                                     submissions=submissions,
                                     existing_grade=existing_grade,
                                     force_makeup=force_makeup)
        
        # 创建或更新评分记录
        if existing_grade:
            existing_grade.grade = grade_float
            existing_grade.feedback = feedback
            existing_grade.is_makeup = is_makeup
            existing_grade.is_cheating = is_cheating
            if is_makeup:
                existing_grade.discount_rate = float(discount_rate)
                existing_grade.original_grade = original_grade
            else:
                existing_grade.discount_rate = None
                existing_grade.original_grade = None
            existing_grade.updated_at = datetime.utcnow()
        else:
            existing_grade = AssignmentGrade(
                assignment_id=assignment_id,
                student_id=student_id,
                teacher_id=current_user.id,
                grade=grade_float,
                feedback=feedback,
                is_makeup=is_makeup,
                is_cheating=is_cheating,
                discount_rate=float(discount_rate) if is_makeup else None,
                original_grade=original_grade if is_makeup else None
            )
            db.session.add(existing_grade)
        
        db.session.commit()
        
        # 记录评分日志
        grade_type = '补交评分' if is_makeup else '正常评分'
        grade_info = f'{grade_float}分' if grade_float is not None else '无评分'
        if is_makeup and original_grade:
            grade_info = f'{original_grade}分×{discount_rate}%={grade_float}分'
        LogService.log_operation(
            operation_type='grade',
            operation_desc=f'{grade_type}：学生 {student.real_name} 的作业「{assignment.title}」，分数：{grade_info}',
            result='success'
        )
        
        current_app.logger.info(f"[SAVE_GRADE] Student={student_id}, Grade={grade_float}, Makeup={is_makeup}, Original={original_grade}, Discount={discount_rate}")
        
        # 创建通知 - 通知学生作业已被评分
        if student.id:
            notification_title = f'作业「{assignment.title}」已被评分'
            notification_content = f'教师 {current_user.real_name} 已对您的作业进行了整体评分'
            if grade_float is not None:
                notification_content += f'，得分：{grade_float}分'
            if feedback:
                notification_content += f'\n\n评语：{feedback[:100]}...' if len(feedback) > 100 else f'\n\n评语：{feedback}'
            
            NotificationService.create_notification(
                sender_id=current_user.id,
                receiver_id=student.id,
                title=notification_title,
                content=notification_content,
                notification_type='grade',
                related_assignment_id=assignment_id
            )
        
        flash(f'已成功给 {student.real_name} 的作业评分')
        
        # 如果是从补交评分入口进入，返回补交评分页面
        if force_makeup:
            return redirect(url_for('assignment.makeup_grading', assignment_id=assignment_id))
        else:
            return redirect(url_for('assignment.view_submissions', assignment_id=assignment_id))
    
    return render_template('grade_assignment_overall.html', 
                         assignment=assignment, 
                         student=student, 
                         submissions=submissions,
                         existing_grade=existing_grade,
                         force_makeup=force_makeup)


@bp.route('/assignment/<int:assignment_id>/student/<int:student_id>/grade', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def grade_student_submissions(assignment_id, student_id):
    """教师给学生的作业进行评分（旧的评分机制 - 针对单次提交）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限评分此作业')
        return redirect(url_for('admin.teacher_dashboard' if current_user.is_teacher else 'admin.super_admin_dashboard'))
    
    # 获取该学生的所有提交记录
    submissions = Submission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).order_by(Submission.submitted_at.desc()).all()
    
    if not submissions:
        flash('该学生尚未提交此作业')
        return redirect(url_for('assignment.view_submissions', assignment_id=assignment_id))
    
    if request.method == 'POST':
        submission_id = request.form.get('submission_id')
        grade = request.form.get('grade')
        feedback = request.form.get('feedback', '')
        
        if not submission_id:
            flash('请选择要评分的提交记录')
            return render_template('grade_student_submissions.html', 
                                 assignment=assignment, 
                                 student=student, 
                                 submissions=submissions)
        
        submission = Submission.query.get_or_404(submission_id)
        
        # 验证评分
        if grade:
            try:
                grade_float = float(grade)
                if grade_float < 0 or grade_float > 100:
                    flash('评分必须在0-100之间')
                    return render_template('grade_student_submissions.html', 
                                         assignment=assignment, 
                                         student=student, 
                                         submissions=submissions)
                submission.grade = grade_float
            except ValueError:
                flash('评分必须是有效的数字')
                return render_template('grade_student_submissions.html', 
                                     assignment=assignment, 
                                     student=student, 
                                     submissions=submissions)
        else:
            submission.grade = None
        
        submission.feedback = feedback
        submission.graded_by = current_user.id
        submission.graded_at = datetime.utcnow()
        
        db.session.commit()
        
        # 创建通知 - 通知学生作业已被批改
        if student.id:  # 确保学生ID存在
            notification_title = f'作业「{assignment.title}」已被批改'
            notification_content = f'教师 {current_user.real_name} 已对您的作业进行了评分'
            if grade:
                notification_content += f'，得分：{grade_float}分'
            if feedback:
                notification_content += f'\n\n评语：{feedback[:100]}...' if len(feedback) > 100 else f'\n\n评语：{feedback}'
            
            NotificationService.create_notification(
                sender_id=current_user.id,
                receiver_id=student.id,
                title=notification_title,
                content=notification_content,
                notification_type='grade',
                related_assignment_id=assignment_id,
                related_submission_id=submission.id
            )
        
        flash(f'已成功评分 {student.real_name} 的作业')
        
        return redirect(url_for('assignment.view_submissions', assignment_id=assignment_id))
    
    return render_template('grade_student_submissions.html', 
                         assignment=assignment, 
                         student=student, 
                         submissions=submissions)


def can_teacher_manage_student(teacher, student):
    """检查教师是否有权限管理某个学生"""
    # 只能管理学生角色
    if student.role != UserRole.STUDENT:
        return False
    
    # 超级管理员可以管理所有学生
    if teacher.is_super_admin:
        return True
    
    # 教师只能管理自己班级中的学生
    if teacher.is_teacher:
        # 获取教师负责的所有班级
        teacher_classes = teacher.teaching_classes
        
        # 检查学生是否在这些班级中
        for class_obj in teacher_classes:
            if student in class_obj.students:
                return True
    
    return False


@bp.route('/assignment/<int:assignment_id>/export_grading_template')
@login_required
@require_teacher_or_admin
def export_grading_template(assignment_id):
    """导出评分模板Excel（仅包含已提交作业的学生）"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        flash('您没有权限导出此作业的评分模板')
        return redirect(url_for('admin.teacher_dashboard' if current_user.is_teacher else 'admin.super_admin_dashboard'))
    
    # 获取所有已提交作业的学生（去重）
    submissions = Submission.query.filter_by(assignment_id=assignment_id).all()
    
    if not submissions:
        flash('该作业暂无学生提交，无法导出评分模板')
        return redirect(url_for('assignment.view_submissions', assignment_id=assignment_id))
    
    # 按学生分组，只保留已提交的学生
    student_map = {}
    for submission in submissions:
        if submission.student_id and submission.student_id not in student_map:
            student_map[submission.student_id] = {
                'student_id': submission.student_id,
                'student_name': submission.student_name,
                'student_number': submission.student_number
            }
    
    # 准备数据
    data = []
    for student_id, student_info in student_map.items():
        # 获取当前教师的评分记录
        grade_record = AssignmentGrade.query.filter_by(
            assignment_id=assignment_id,
            student_id=student_id,
            teacher_id=current_user.id
        ).first()
        
        # 确定状态
        status = '正常'
        if grade_record and grade_record.is_cheating:
            status = '作弊/抄袭'
        
        row = {
            '学号': student_info['student_number'],
            '姓名': student_info['student_name'],
            '分数': grade_record.grade if grade_record and grade_record.grade is not None else '',
            '评语': grade_record.feedback if grade_record and grade_record.feedback else '',
            '状态': status
        }
        data.append(row)
    
    # 按学号排序
    data.sort(key=lambda x: x['学号'] if x['学号'] else '')
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='评分表', index=False)
        
        # 获取worksheet进行格式调整
        workbook = writer.book
        worksheet = writer.sheets['评分表']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 15  # 学号
        worksheet.column_dimensions['B'].width = 12  # 姓名
        worksheet.column_dimensions['C'].width = 10  # 分数
        worksheet.column_dimensions['D'].width = 50  # 评语
        worksheet.column_dimensions['E'].width = 15  # 状态
        
        # 设置表头样式
        from openpyxl.styles import Font, Alignment, PatternFill
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置数据居中（学号、姓名、分数、状态）
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for i, cell in enumerate(row):
                if i < 3:  # 学号、姓名、分数
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif i == 4:  # 状态
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # 评语
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 冻结第一行
        worksheet.freeze_panes = 'A2'
        
        # 为状态列添加数据验证（下拉列表）
        from openpyxl.worksheet.datavalidation import DataValidation
        
        dv = DataValidation(type="list", formula1='"正常,作弊/抄袭"', allow_blank=False)
        dv.error = '请从下拉列表中选择状态'
        dv.errorTitle = '输入错误'
        dv.prompt = '请选择：正常 或 作弊/抄袭'
        dv.promptTitle = '状态选择'
        worksheet.add_data_validation(dv)
        # 应用到所有状态列单元格（E2开始到最后一行）
        dv.add(f'E2:E{worksheet.max_row}')
        
        # 添加说明sheet
        instructions_df = pd.DataFrame({
            '说明': [
                '1. 请仅修改「分数」、「评语」和「状态」列',
                '2. 分数范围：0-100',
                '3. 分数可以为空，评语可选',
                '4. 状态只能是“正常”或“作弊/抄袭”，请从下拉列表中选择',
                '5. 标记为“作弊/抄袭”的作业，无论多少分，成绩统计中一律为0分',
                '6. 请勿修改学号和姓名',
                '7. 请勿删除或添加行',
                '8. 填写完成后，保存并上传此文件'
            ]
        })
        instructions_df.to_excel(writer, sheet_name='使用说明', index=False)
        worksheet2 = writer.sheets['使用说明']
        worksheet2.column_dimensions['A'].width = 60
        for cell in worksheet2['A']:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    output.seek(0)
    
    # 生成文件名
    beijing_time = to_beijing_time(datetime.utcnow())
    timestamp = beijing_time.strftime('%Y%m%d_%H%M%S')
    safe_title = assignment.title.replace('/', '_').replace('\\', '_')[:50]
    filename = f'{safe_title}_评分模板_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/assignment/<int:assignment_id>/import_grades', methods=['POST'])
@login_required
@require_teacher_or_admin
def import_grades(assignment_id):
    """批量导入评分"""
    assignment = Assignment.query.get_or_404(assignment_id)
    
    # 权限检查
    if not can_manage_assignment(current_user, assignment):
        return jsonify({'success': False, 'message': '您没有权限导入此作业的评分'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择Excel文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择Excel文件'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '只支持Excel文件(.xlsx或.xls)'}), 400
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file, sheet_name='评分表')
        
        # 验证必要列
        required_columns = ['学号', '姓名', '分数', '评语']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'Excel文件缺少必要列：{", ".join(missing_columns)}'
            }), 400
        
        success_count = 0
        error_count = 0
        errors = []
        updated_students = []  # 记录被更新评分的学生
        
        # 获取所有已提交作业的学生ID集合
        submitted_student_ids = set()
        submissions = Submission.query.filter_by(assignment_id=assignment_id).all()
        for submission in submissions:
            if submission.student_id:
                submitted_student_ids.add(submission.student_id)
        
        for index, row in df.iterrows():
            try:
                row_num = index + 2  # Excel行号（从2开始，因为第1行是表头）
                
                student_number = str(row['学号']).strip() if pd.notna(row['学号']) else ''
                student_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else ''
                grade_str = str(row['分数']).strip() if pd.notna(row['分数']) else ''
                feedback = str(row['评语']).strip() if pd.notna(row['评语']) else ''
                status = str(row['状态']).strip() if '状态' in row and pd.notna(row['状态']) else '正常'
                
                if not student_number or not student_name:
                    errors.append(f'第{row_num}行：学号或姓名为空')
                    error_count += 1
                    continue
                
                # 查找学生
                student = User.query.filter_by(
                    student_id=student_number,
                    real_name=student_name,
                    role=UserRole.STUDENT
                ).first()
                
                if not student:
                    errors.append(f'第{row_num}行：找不到学号为「{student_number}」姓名为「{student_name}」的学生')
                    error_count += 1
                    continue
                
                # 检查学生是否已提交作业
                if student.id not in submitted_student_ids:
                    errors.append(f'第{row_num}行：学生「{student_name}」未提交此作业，无法评分')
                    error_count += 1
                    continue
                
                # 解析分数
                grade_float = None
                if grade_str:
                    try:
                        grade_float = float(grade_str)
                        if grade_float < 0 or grade_float > 100:
                            errors.append(f'第{row_num}行：分数必须在0-100之间')
                            error_count += 1
                            continue
                    except ValueError:
                        errors.append(f'第{row_num}行：分数格式错误')
                        error_count += 1
                        continue
                
                # 验证状态
                is_cheating = False
                if status not in ['正常', '作弊/抄袭']:
                    errors.append(f'第{row_num}行：状态只能是“正常”或“作弊/抄袭”')
                    error_count += 1
                    continue
                
                if status == '作弊/抄袭':
                    is_cheating = True
                
                # 查找或创建评分记录
                grade_record = AssignmentGrade.query.filter_by(
                    assignment_id=assignment_id,
                    student_id=student.id,
                    teacher_id=current_user.id
                ).first()
                
                if grade_record:
                    # 更新现有记录
                    grade_record.grade = grade_float
                    grade_record.feedback = feedback
                    grade_record.is_cheating = is_cheating
                    grade_record.updated_at = datetime.utcnow()
                else:
                    # 创建新记录
                    grade_record = AssignmentGrade(
                        assignment_id=assignment_id,
                        student_id=student.id,
                        teacher_id=current_user.id,
                        grade=grade_float,
                        feedback=feedback,
                        is_cheating=is_cheating
                    )
                    db.session.add(grade_record)
                
                success_count += 1
                updated_students.append(student)
                
            except Exception as e:
                error_msg = f'第{row_num}行：处理失败 - {str(e)}'
                errors.append(error_msg)
                error_count += 1
                current_app.logger.error(error_msg)
        
        # 提交所有更改
        db.session.commit()
        
        # 发送通知给被评分的学生
        for student in updated_students:
            try:
                grade_record = AssignmentGrade.query.filter_by(
                    assignment_id=assignment_id,
                    student_id=student.id,
                    teacher_id=current_user.id
                ).first()
                
                notification_title = f'作业「{assignment.title}」已被评分'
                notification_content = f'教师 {current_user.real_name} 已对您的作业进行了评分'
                if grade_record.grade is not None:
                    notification_content += f'，得分：{grade_record.grade}分'
                if grade_record.feedback:
                    feedback_preview = grade_record.feedback[:100] + '...' if len(grade_record.feedback) > 100 else grade_record.feedback
                    notification_content += f'\n\n评语：{feedback_preview}'
                
                NotificationService.create_notification(
                    sender_id=current_user.id,
                    receiver_id=student.id,
                    title=notification_title,
                    content=notification_content,
                    notification_type='grade',
                    related_assignment_id=assignment_id
                )
            except Exception as e:
                current_app.logger.error(f'发送通知失败: {str(e)}')
        
        # 构建返回消息
        message = f'导入完成：成功 {success_count} 条，失败 {error_count} 条'
        if errors and len(errors) <= 10:
            message += '<br><br>错误详情：<br>' + '<br>'.join(errors)
        elif errors:
            message += f'<br><br>错误详情（前10条）：<br>' + '<br>'.join(errors[:10])
            message += f'<br>...还有 {len(errors) - 10} 条错误未显示'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'批量导入评分失败: {str(e)}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'导入失败：{str(e)}'
        }), 500
