"""班级管理路由"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
import os
from io import BytesIO
from datetime import datetime
import pandas as pd
from app.extensions import db
from app.models import User, UserRole, Class, Assignment
from app.utils import require_teacher_or_admin, require_role, to_beijing_time

bp = Blueprint('class_mgmt', __name__, url_prefix='/admin/classes')


@bp.route('/')
@login_required
@require_teacher_or_admin
def manage_classes():
    """班级管理列表"""
    if current_user.is_super_admin:
        classes = Class.query.order_by(Class.created_at.desc()).all()
    else:
        # 教师只能看到自己的班级
        classes = current_user.teaching_classes
    
    # 计算统计数据（去重）
    if classes:
        # 收集所有唯一的学生ID
        all_student_ids = set()
        # 收集所有唯一的教师ID
        all_teacher_ids = set()
        # 收集所有作业ID（作业不会重复，因为每个作业只属于一个班级）
        total_assignments = 0
        
        for class_obj in classes:
            # 添加学生ID到集合（自动去重）
            all_student_ids.update([s.id for s in class_obj.students])
            # 添加教师ID到集合（自动去重）
            all_teacher_ids.update([t.id for t in class_obj.teachers])
            # 累加作业数
            total_assignments += len(class_obj.assignments)
        
        # 统计数据
        stats = {
            'total_students': len(all_student_ids),
            'total_teachers': len(all_teacher_ids),
            'total_assignments': total_assignments
        }
    else:
        stats = {
            'total_students': 0,
            'total_teachers': 0,
            'total_assignments': 0
        }
    
    return render_template('manage_classes.html', classes=classes, stats=stats)


@bp.route('/add', methods=['GET', 'POST'])
@login_required
@require_role(UserRole.SUPER_ADMIN)
def add_class():
    """创建班级"""
    if request.method == 'POST':
        name = request.form['name']
        code = request.form['code']
        description = request.form.get('description', '')
        grade = request.form.get('grade', '')
        
        # 检查班级代码是否已存在
        if Class.query.filter_by(code=code).first():
            flash('班级代码已存在')
            return render_template('add_class.html')
        
        # 创建新班级
        new_class = Class(
            name=name,
            code=code,
            description=description,
            grade=grade,
            created_by=current_user.id
        )
        
        db.session.add(new_class)
        db.session.commit()
        
        flash(f'班级 {name} 创建成功！现在可以在创建用户时将教师和学生分配到该班级')
        return redirect(url_for('class_mgmt.manage_classes'))
    
    return render_template('add_class.html')


@bp.route('/<int:class_id>/edit', methods=['GET', 'POST'])
@login_required
@require_teacher_or_admin
def edit_class(class_id):
    """编辑班级"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限编辑此班级')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    if request.method == 'POST':
        class_obj.name = request.form['name']
        class_obj.code = request.form['code']
        class_obj.description = request.form.get('description', '')
        class_obj.grade = request.form.get('grade', '')
        class_obj.is_active = 'is_active' in request.form
        
        # 更新教师分配（只有超级管理员可以修改）
        if current_user.is_super_admin:
            teacher_ids = request.form.getlist('teachers')
            class_obj.teachers.clear()
            if teacher_ids:
                teachers = User.query.filter(
                    User.id.in_(teacher_ids),
                    User.role == UserRole.TEACHER
                ).all()
                class_obj.teachers.extend(teachers)
        
        db.session.commit()
        flash(f'班级 {class_obj.name} 信息已更新')
        return redirect(url_for('class_mgmt.manage_classes'))
    
    teachers = User.query.filter_by(role=UserRole.TEACHER).all()
    return render_template('edit_class.html', class_obj=class_obj, teachers=teachers)


@bp.route('/<int:class_id>/delete', methods=['POST'])
@login_required
@require_teacher_or_admin
def delete_class(class_id):
    """删除班级"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        # 教师只能删除自己负责的空班级
        if current_user not in class_obj.teachers:
            flash('您没有权限删除此班级')
            return redirect(url_for('class_mgmt.manage_classes'))
        
        # 检查班级是否为空
        if class_obj.students or class_obj.assignments:
            flash('只能删除没有学生和作业的空班级')
            return redirect(url_for('class_mgmt.manage_classes'))
    else:
        # 超级管理员删除班级时的清理工作
        class_obj.students.clear()
        class_obj.teachers.clear()
        
        # 删除相关作业和提交文件
        for assignment in class_obj.assignments:
            for submission in assignment.submissions:
                try:
                    if os.path.exists(submission.file_path):
                        os.remove(submission.file_path)
                except Exception as e:
                    print(f"删除文件失败: {e}")
    
    class_name = class_obj.name
    db.session.delete(class_obj)
    db.session.commit()
    
    flash(f'班级 "{class_name}" 已成功删除')
    return redirect(url_for('class_mgmt.manage_classes'))


@bp.route('/<int:class_id>/students')
@login_required
@require_teacher_or_admin
def class_students(class_id):
    """班级学生管理"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    # 获取可以添加到班级的学生
    current_student_ids = [student.id for student in class_obj.students]
    available_students = User.query.filter(
        User.role == UserRole.STUDENT,
        ~User.id.in_(current_student_ids) if current_student_ids else True
    ).order_by(User.real_name).all()
    
    return render_template('class_students.html',
                         class_obj=class_obj,
                         available_students=available_students)


@bp.route('/<int:class_id>/add_student', methods=['POST'])
@login_required
@require_teacher_or_admin
def add_student_to_class(class_id):
    """添加学生到班级"""
    class_obj = Class.query.get_or_404(class_id)
    student_id = request.form.get('student_id')
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    student = User.query.filter_by(id=student_id, role=UserRole.STUDENT).first()
    if not student:
        flash('学生不存在')
        return redirect(url_for('class_mgmt.class_students', class_id=class_id))
    
    if student in class_obj.students:
        flash(f'{student.real_name} 已经在班级中')
    else:
        class_obj.students.append(student)
        db.session.commit()
        flash(f'已将 {student.real_name} 添加到班级')
    
    return redirect(url_for('class_mgmt.class_students', class_id=class_id))


@bp.route('/<int:class_id>/remove_student', methods=['POST'])
@login_required
@require_teacher_or_admin
def remove_student_from_class(class_id):
    """从班级移除学生"""
    class_obj = Class.query.get_or_404(class_id)
    student_id = request.form.get('student_id')
    student = User.query.get_or_404(student_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限管理此班级')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    if student in class_obj.students:
        class_obj.students.remove(student)
        db.session.commit()
        flash(f'已将 {student.real_name} 从班级中移除')
    else:
        flash(f'{student.real_name} 不在此班级中')
    
    return redirect(url_for('class_mgmt.class_students', class_id=class_id))


@bp.route('/<int:class_id>/grades')
@login_required
@require_teacher_or_admin
def class_grades(class_id):
    """班级成绩统计"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限查看此班级成绩')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    # 获取班级所有作业
    assignments = Assignment.query.filter_by(class_id=class_id).order_by(Assignment.created_at).all()
    
    # 获取班级所有学生
    students = class_obj.students
    
    # 构建成绩统计数据
    from app.models import Submission, AssignmentGrade
    
    grade_stats = []
    for student in students:
        student_data = {
            'student': student,
            'grades': {},
            'graded_count': 0,
            'total_score': 0,  # 新增：总分
            'average': 0
        }
        
        total_score = 0  # 所有作业的总分（未交作业计0分）
        graded_assignments = 0
        
        for assignment in assignments:
            # 首先检查 AssignmentGrade 表（支持补交评分）
            avg_grade = get_student_assignment_average_grade(assignment.id, student.id)
            
            # 检查是否为补交作业
            grade_record = AssignmentGrade.query.filter_by(
                assignment_id=assignment.id,
                student_id=student.id
            ).first()
            is_makeup = grade_record.is_makeup if grade_record and grade_record.is_makeup else False
            original_grade = grade_record.original_grade if grade_record and grade_record.original_grade else None
            discount_rate = grade_record.discount_rate if grade_record and grade_record.discount_rate else None
            
            if avg_grade is not None:
                # 有评分记录（可能是补交评分）
                student_data['grades'][assignment.id] = {
                    'grade': avg_grade, 
                    'status': 'graded', 
                    'is_makeup': is_makeup,
                    'original_grade': original_grade,
                    'discount_rate': discount_rate
                }
                total_score += avg_grade
                graded_assignments += 1
            else:
                # 没有评分记录，检查是否有提交记录
                has_submission = Submission.query.filter_by(
                    assignment_id=assignment.id,
                    student_id=student.id
                ).first()
                
                if not has_submission:
                    # 未提交作业：记为0分并标记“未交”
                    student_data['grades'][assignment.id] = {'grade': 0, 'status': 'not_submitted', 'is_makeup': False}
                    total_score += 0
                else:
                    # 已提交但未评分，尝试从旧系统获取
                    submission = Submission.query.filter_by(
                        assignment_id=assignment.id,
                        student_id=student.id
                    ).filter(Submission.grade.isnot(None)).order_by(
                        Submission.graded_at.desc()
                    ).first()
                    
                    if submission and submission.grade is not None:
                        student_data['grades'][assignment.id] = {'grade': submission.grade, 'status': 'graded', 'is_makeup': False}
                        total_score += submission.grade
                        graded_assignments += 1
                    else:
                        # 已提交但未评分
                        student_data['grades'][assignment.id] = {'grade': 0, 'status': 'submitted_not_graded', 'is_makeup': False}
                        total_score += 0
        
        # 设置总分
        student_data['total_score'] = round(total_score, 2)
        student_data['graded_count'] = graded_assignments
        
        # 计算平均分：总分除以作业总次数（包括未交作业）
        total_assignments = len(assignments)
        if total_assignments > 0:
            student_data['average'] = round(total_score / total_assignments, 2)
        else:
            student_data['average'] = 0
        
        grade_stats.append(student_data)
    
    # 按平均分排序
    grade_stats.sort(key=lambda x: x['average'], reverse=True)
    
    # 添加排名
    for rank, student_data in enumerate(grade_stats, 1):
        student_data['rank'] = rank
    
    return render_template('class_grades.html',
                         class_obj=class_obj,
                         assignments=assignments,
                         grade_stats=grade_stats)


def get_student_assignment_average_grade(assignment_id, student_id):
    """获取学生在某作业的平均分"""
    from app.models import AssignmentGrade
    from sqlalchemy import func
    
    avg_grade = db.session.query(func.avg(AssignmentGrade.grade)).filter(
        AssignmentGrade.assignment_id == assignment_id,
        AssignmentGrade.student_id == student_id
    ).scalar()
    
    return round(avg_grade, 2) if avg_grade is not None else None


@bp.route('/<int:class_id>/export_grades')
@login_required
@require_teacher_or_admin
def export_class_grades(class_id):
    """导出班级成绩为Excel"""
    class_obj = Class.query.get_or_404(class_id)
    
    # 权限检查
    if not current_user.is_super_admin:
        if current_user not in class_obj.teachers:
            flash('您没有权限导出此班级成绩')
            return redirect(url_for('class_mgmt.manage_classes'))
    
    # 获取班级所有作业
    assignments = Assignment.query.filter_by(class_id=class_id).order_by(Assignment.created_at).all()
    
    # 获取班级所有学生
    students = class_obj.students
    
    if not students or not assignments:
        flash('班级暂无学生或作业，无法导出')
        return redirect(url_for('class_mgmt.class_grades', class_id=class_id))
    
    # 构建成绩数据
    from app.models import Submission, AssignmentGrade
    
    # 准备数据
    data = []
    for student in students:
        row = {
            '排名': 0,  # 稍后填充
            '姓名': student.real_name,
            '学号': student.student_id or student.username,
        }
        
        total_score = 0
        graded_count = 0
        
        # 每个作业的成绩
        for assignment in assignments:
            # 首先检查 AssignmentGrade 表（支持补交评分）
            avg_grade = get_student_assignment_average_grade(assignment.id, student.id)
            
            if avg_grade is not None:
                # 有评分记录（可能是补交评分）
                row[assignment.title] = avg_grade
                total_score += avg_grade
                graded_count += 1
            else:
                # 没有评分记录，检查是否有提交记录
                has_submission = Submission.query.filter_by(
                    assignment_id=assignment.id,
                    student_id=student.id
                ).first()
                
                if not has_submission:
                    # 未提交作业：记为0分并标记"未交"
                    row[assignment.title] = '0分(未交)'
                    total_score += 0
                else:
                    # 已提交作业：获取评分
                    avg_grade = get_student_assignment_average_grade(assignment.id, student.id)
                    
                    if avg_grade is not None:
                        row[assignment.title] = avg_grade
                        total_score += avg_grade
                        graded_count += 1
                    else:
                        # 尝试从旧系统获取
                        submission = Submission.query.filter_by(
                            assignment_id=assignment.id,
                            student_id=student.id
                        ).filter(Submission.grade.isnot(None)).order_by(
                            Submission.graded_at.desc()
                        ).first()
                        
                        if submission and submission.grade is not None:
                            row[assignment.title] = submission.grade
                            total_score += submission.grade
                            graded_count += 1
                        else:
                            # 已提交但未评分
                            row[assignment.title] = '未评分'
                            total_score += 0

        # 总分和平均分
        row['总分'] = round(total_score, 2)
        total_assignments = len(assignments)
        row['平均分'] = round(total_score / total_assignments, 2) if total_assignments > 0 else 0
        row['评分进度'] = f'{graded_count}/{total_assignments}'
        
        data.append(row)
    
    # 按平均分排序
    data.sort(key=lambda x: x['平均分'], reverse=True)
    
    # 填充排名
    for rank, row in enumerate(data, 1):
        row['排名'] = rank
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 调整列顺序
    columns = ['排名', '姓名', '学号'] + [a.title for a in assignments] + ['总分', '平均分', '评分进度']
    df = df[columns]
    
    # 创建Excel文件
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='成绩统计', index=False)
        
        # 获取worksheet进行格式调整
        workbook = writer.book
        worksheet = writer.sheets['成绩统计']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 8   # 排名
        worksheet.column_dimensions['B'].width = 12  # 姓名
        worksheet.column_dimensions['C'].width = 15  # 学号
        
        # 作业列宽
        for i, assignment in enumerate(assignments, start=4):
            col_letter = chr(64 + i)  # D, E, F...
            worksheet.column_dimensions[col_letter].width = 15
        
        # 总分、平均分、评分进度列宽
        last_col_index = 4 + len(assignments)
        worksheet.column_dimensions[chr(64 + last_col_index)].width = 12      # 总分
        worksheet.column_dimensions[chr(64 + last_col_index + 1)].width = 12  # 平均分
        worksheet.column_dimensions[chr(64 + last_col_index + 2)].width = 15  # 评分进度
        
        # 设置表头样式
        from openpyxl.styles import Font, Alignment, PatternFill
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置数据居中
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 冻结第一行
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    
    # 生成文件名
    beijing_time = to_beijing_time(datetime.utcnow())
    timestamp = beijing_time.strftime('%Y%m%d_%H%M%S')
    filename = f'{class_obj.name}_成绩统计_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
    beijing_time = to_beijing_time(datetime.utcnow())
    timestamp = beijing_time.strftime('%Y%m%d_%H%M%S')
    filename = f'{class_obj.name}_成绩统计_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
